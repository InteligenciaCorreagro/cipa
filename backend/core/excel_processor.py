# src/excel_processor.py

from openpyxl import Workbook
from openpyxl.styles import numbers, Font, PatternFill, Alignment
from datetime import datetime, timedelta
from typing import List, Dict
import logging
import re

logger = logging.getLogger(__name__)

class ExcelProcessor:
    """Procesa y genera archivos Excel con datos de facturas"""
    
    # Constantes de la empresa
    NIT_VENDEDOR = "890907163"
    NOMBRE_VENDEDOR = "COMPAÑIA INDUSTRIAL DE PRODUCTOS AGROPECUARIOS S.A"
    CODIGO_SUBYACENTE = "SPN-1"
    
    def __init__(self, template_path: str = None):
        self.template_path = template_path
    
    def _extraer_dias_condicion_pago(self, condicion_pago: str) -> int:
        """
        Extrae el número de días de la condición de pago
        
        Ejemplos:
        - "CREDITO 15 DIAS" → 15
        - "CREDITO 8 DIAS" → 8
        - "CREDITO 30 DIAS" → 30
        - "CONTADO" → 0
        - "15 DIAS" → 15
        
        Args:
            condicion_pago: Descripción de la condición de pago
            
        Returns:
            Número de días (0 si es contado o no se encuentra)
        """
        if not condicion_pago:
            return 0
        
        condicion_upper = str(condicion_pago).upper().strip()
        
        # Si es contado, retornar 0
        if 'CONTADO' in condicion_upper:
            return 0
        
        # Buscar patrón de número seguido de "DIAS" o "DIA"
        # Ejemplos: "15 DIAS", "CREDITO 30 DIAS", "8 DIA"
        match = re.search(r'(\d+)\s*DIA[S]?', condicion_upper)
        if match:
            dias = int(match.group(1))
            logger.debug(f"Extraídos {dias} días de condición: '{condicion_pago}'")
            return dias
        
        # Si no se encuentra patrón, buscar cualquier número
        match = re.search(r'(\d+)', condicion_upper)
        if match:
            dias = int(match.group(1))
            logger.debug(f"Extraído número {dias} de condición: '{condicion_pago}'")
            return dias
        
        # Por defecto, retornar 0 (pago inmediato)
        logger.debug(f"No se encontraron días en condición: '{condicion_pago}', usando 0")
        return 0
    
    def _calcular_fecha_pago(self, fecha_factura, condicion_pago: str):
        """
        Calcula la fecha de pago sumando los días de la condición de pago a la fecha de factura
        
        Args:
            fecha_factura: Fecha de la factura (datetime.date o datetime)
            condicion_pago: Descripción de la condición de pago
            
        Returns:
            Fecha de pago (datetime.date) o None si no hay fecha de factura
        """
        if not fecha_factura:
            return None
        
        # Asegurar que es un objeto date
        if isinstance(fecha_factura, datetime):
            fecha_factura = fecha_factura.date()
        
        # Extraer días de la condición de pago
        dias = self._extraer_dias_condicion_pago(condicion_pago)
        
        # Calcular fecha de pago
        fecha_pago = fecha_factura + timedelta(days=dias)
        
        logger.debug(f"Fecha factura: {fecha_factura}, Condición: '{condicion_pago}' ({dias} días) → Fecha pago: {fecha_pago}")
        
        return fecha_pago
    
    def transformar_factura(self, factura: Dict) -> Dict:
        """
        Transforma una factura del formato API al formato Excel

        Args:
            factura: Datos de factura en formato JSON de la API

        Returns:
            Diccionario con datos transformados
        """
        # N° Factura: Prefijo + Número (sin espacios)
        prefijo = str(factura.get('f_prefijo', '')).strip()
        nrodocto = factura.get('f_nrodocto', '')
        numero_factura = f"{prefijo}{nrodocto}"

        # Procesar fecha factura
        fecha_str = factura.get('f_fecha', '')
        fecha_factura = None
        if fecha_str:
            try:
                fecha_factura = datetime.fromisoformat(str(fecha_str).replace('T00:00:00', ''))
            except (ValueError, AttributeError):
                logger.warning(f"Error parseando fecha: {fecha_str}")

        # Calcular fecha de pago
        condicion_pago = factura.get('f_desc_cond_pago', '')
        fecha_pago = self._calcular_fecha_pago(fecha_factura, condicion_pago)

        # Extraer IVA del grupo impositivo (solo el número)
        iva = self._extraer_iva(factura.get('f_desc_grupo_impositivo', ''))

        # Extraer ciudad (sin el código)
        ciudad = self._extraer_ciudad(factura.get('f_ciudad_punto_envio'))

        # Normalizar unidad de medida a KG, UN o LT
        unidad_medida = self._normalizar_unidad_medida(factura.get('f_um_inv_desc', ''))

        # Cantidad base original de la API
        cantidad_base_api = factura.get('f_cant_base', 0.0)
        if cantidad_base_api is None:
            cantidad_base_api = 0.0

        # f_um_base original (para la última columna)
        um_base = str(factura.get('f_um_base', '')).strip()

        # Multiplicador de unidad base
        multiplicador = self._extraer_multiplicador_um_base(um_base)

        # INTERCAMBIO: 
        # - cantidad (columna E) = cantidad_base_api * multiplicador (lo que antes era cantidad_original)
        # - cantidad_original (columna T) = cantidad_base_api (lo que antes era cantidad)
        cantidad_convertida = float(cantidad_base_api) * multiplicador
        cantidad_original = float(cantidad_base_api)

        # Valor total del API
        valor_total = factura.get('f_valor_subtotal_local', 0.0)
        if valor_total is None:
            valor_total = 0.0
        valor_total = float(valor_total)

        # Calcular precio unitario = valor_total / cantidad_convertida
        if cantidad_convertida != 0:
            precio_unitario = valor_total / cantidad_convertida
        else:
            precio_unitario = 0.0
            logger.warning(f"Cantidad convertida es 0 para factura {numero_factura}, precio unitario = 0")

        # Código de producto: usar f_cod_item si existe, sino f_desc_item (igual que las notas)
        codigo_producto_raw = factura.get('f_cod_item') or factura.get('f_desc_item', '')
        codigo_producto = str(codigo_producto_raw).strip()

        return {
            'numero_factura': numero_factura,
            'nombre_producto': str(factura.get('f_desc_item', '')).strip(),
            'codigo_subyacente': self.CODIGO_SUBYACENTE,
            'unidad_medida': unidad_medida,
            'cantidad': cantidad_convertida,
            'precio_unitario': precio_unitario,
            'fecha_factura': fecha_factura,
            'fecha_pago': fecha_pago,  # AHORA CALCULADA
            'nit_comprador': str(factura.get('f_cliente_desp', '')).strip(),
            'nombre_comprador': str(factura.get('f_cliente_fact_razon_soc', '')).strip(),
            'nit_vendedor': self.NIT_VENDEDOR,
            'nombre_vendedor': self.NOMBRE_VENDEDOR,
            'principal': 'V',
            'municipio': ciudad,
            'iva': iva,
            'descripcion': str(factura.get('f_desc_tipo_inv', '')).strip(),
            'activa_factura': '1',
            'activa_bodega': '1',
            'incentivo': '',
            'cantidad_original': cantidad_original,
            'moneda': '1',
            'um_base': um_base,
            'valor_total': valor_total,
            'codigo_producto_api': codigo_producto,  # Ahora usa f_cod_item o f_desc_item (igual que notas)
            'condicion_pago': condicion_pago  # Para debug/auditoría
        }
    
    def _extraer_iva(self, grupo_impositivo: str) -> str:
        """
        Extrae el porcentaje de IVA del grupo impositivo
        Ejemplo: "IVA 5% RTF BIENES RETEIVA RETEICA" -> "5"
        """
        if not grupo_impositivo:
            return '0'
        
        grupo_impositivo = str(grupo_impositivo)
        
        # Buscar patrón "IVA X%" o "X%"
        match = re.search(r'IVA\s*(\d+)%', grupo_impositivo, re.IGNORECASE)
        if match:
            return match.group(1)
        
        # Si no encuentra, buscar solo número con %
        match = re.search(r'(\d+)%', grupo_impositivo)
        if match:
            return match.group(1)
        
        # Por defecto 0
        return '0'
    
    def _extraer_ciudad(self, ciudad_raw) -> str:
        """
        Extrae el nombre de la ciudad del formato 'XXX-Ciudad'
        Ejemplo: "001-Pereira" -> "Pereira"
        """
        # Manejar valores None o vacíos
        if ciudad_raw is None or ciudad_raw == '':
            return ''
        
        ciudad_raw = str(ciudad_raw).strip()
        
        if '-' in ciudad_raw:
            return ciudad_raw.split('-', 1)[1].strip()
        
        return ciudad_raw
    
    def _normalizar_unidad_medida(self, um_desc: str) -> str:
        """
        Normaliza la unidad de medida a KG, UN o LT
        """
        if not um_desc:
            return 'UN'
        
        um_upper = str(um_desc).upper().strip()
        
        # Mapeo de unidades - IMPORTANTE: Evaluar términos más específicos primero
        # BULTO debe evaluarse antes que UNIDAD porque contiene 'UN'
        if 'BULTO' in um_upper or 'BT' in um_upper:
            return 'KG'  # Los bultos se consideran en KG
        elif 'KILO' in um_upper or 'KG' in um_upper or 'KLS' in um_upper:
            return 'KG'
        elif 'LITRO' in um_upper or 'LT' in um_upper:
            return 'LT'
        elif 'UNIDAD' in um_upper or 'UND' in um_upper or 'UN' in um_upper:
            return 'UN'
        else:
            # Por defecto UN
            return 'UN'
    
    def _extraer_multiplicador_um_base(self, um_base: str) -> float:
        """
        Extrae el multiplicador de la unidad base
        Ejemplos:
        - "BT40" -> 40
        - "BT30" -> 30
        - "KLS" -> 1
        - "UND" -> 1
        - "800G" -> 0.8 (gramos a kilogramos)
        """
        if not um_base:
            return 1.0

        # Buscar número en la cadena
        match = re.search(r'(\d+)', str(um_base))
        if match:
            numero = float(match.group(1))
            # Si termina en "G" (gramos), convertir a kilogramos dividiendo entre 1000
            if um_base.upper().endswith('G'):
                return numero / 1000.0
            return numero

        # Si no tiene número, multiplicar por 1
        return 1.0
    
    def _crear_encabezados(self, ws):
        """Crea los encabezados del Excel con formato"""
        encabezados = [
            'N° Factura',
            'Nombre Producto',
            'Codigo Subyacente',
            'Unidad Medida en Kg,Un,Lt',
            'Cantidad (5 decimales - separdor coma)',
            'Precio Unitario (5 decimales - separdor coma)',
            'Fecha Factura Año-Mes-Dia',
            'Fecha Pago Año-Mes-Dia',
            'Nit Comprador (Existente)',
            'Nombre Comprador',
            'Nit Vendedor (Existente)',
            'Nombre Vendedor',
            'Principal V,C',
            'Municipio (Nombre Exacto de la Ciudad)',
            'Iva (N°%)',
            'Descripción',
            'Activa Factura',
            'Activa Bodega',
            'Incentivo',
            'Cantidad Original (5 decimales - separdor coma)',
            'Moneda (1,2,3)',
            'UM Base',
            'Valor Total'
        ]

        # Estilo de encabezado
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        for col_num, header in enumerate(encabezados, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        # Ajustar ancho de columnas
        column_widths = [15, 40, 18, 25, 25, 25, 22, 22, 22, 40, 22, 50, 15, 35, 12, 35, 15, 15, 15, 30, 15, 15, 20]
        for col_num, width in enumerate(column_widths, 1):
            ws.column_dimensions[ws.cell(row=1, column=col_num).column_letter].width = width
    
    def generar_excel(self, facturas: List[Dict], output_path: str) -> str:
        """
        Genera archivo Excel con las facturas
        
        Args:
            facturas: Lista de facturas transformadas
            output_path: Ruta donde guardar el archivo
            
        Returns:
            Ruta del archivo generado
        """
        try:
            # Crear nuevo workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Facturas"
            
            # Crear encabezados
            self._crear_encabezados(ws)
            
            logger.info(f"Procesando {len(facturas)} facturas")
            
            # Agregar datos
            for idx, factura in enumerate(facturas, start=2):
                ws[f'A{idx}'] = factura['numero_factura']
                ws[f'B{idx}'] = factura['nombre_producto']
                ws[f'C{idx}'] = factura['codigo_subyacente']
                ws[f'D{idx}'] = factura['unidad_medida']
                
                # Cantidad con 5 decimales y separador coma
                ws[f'E{idx}'] = factura['cantidad']
                ws[f'E{idx}'].number_format = '#,##0.00000'
                
                # Precio unitario con 5 decimales
                ws[f'F{idx}'] = factura['precio_unitario']
                ws[f'F{idx}'].number_format = '#,##0.00000'
                
                # Fecha factura
                if factura['fecha_factura']:
                    ws[f'G{idx}'] = factura['fecha_factura']
                    ws[f'G{idx}'].number_format = 'YYYY-MM-DD'
                
                # Fecha pago (AHORA CALCULADA)
                if factura['fecha_pago']:
                    ws[f'H{idx}'] = factura['fecha_pago']
                    ws[f'H{idx}'].number_format = 'YYYY-MM-DD'
                else:
                    ws[f'H{idx}'] = ''
                
                ws[f'I{idx}'] = factura['nit_comprador']
                ws[f'J{idx}'] = factura['nombre_comprador']
                ws[f'K{idx}'] = factura['nit_vendedor']
                ws[f'L{idx}'] = factura['nombre_vendedor']
                ws[f'M{idx}'] = factura['principal']
                ws[f'N{idx}'] = factura['municipio']
                ws[f'O{idx}'] = factura['iva']
                ws[f'P{idx}'] = factura['descripcion']
                ws[f'Q{idx}'] = factura['activa_factura']
                ws[f'R{idx}'] = factura['activa_bodega']
                ws[f'S{idx}'] = factura['incentivo']
                
                # Cantidad original con 5 decimales
                ws[f'T{idx}'] = factura['cantidad_original']
                ws[f'T{idx}'].number_format = '#,##0.00000'
                
                ws[f'U{idx}'] = factura['moneda']
                ws[f'V{idx}'] = factura['um_base']
                
                # Valor Total con 5 decimales
                ws[f'W{idx}'] = factura['valor_total']
                ws[f'W{idx}'].number_format = '#,##0.00000'
            
            # Guardar archivo
            wb.save(output_path)
            logger.info(f"Excel generado exitosamente: {output_path}")
            return output_path
            
        except Exception as e:
            logger.error(f"Error al generar Excel: {e}")
            raise