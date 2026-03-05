"""
API REST para consulta de notas de crédito - VERSIÓN REESTRUCTURADA

Endpoints para:
- Autenticación de usuarios
- Consulta de facturas, notas crédito y facturas rechazadas
- Dashboard y reportes
"""

import os
import sys
import logging
import time
import json
from datetime import datetime, timedelta
from pathlib import Path
from flask import Flask, request, jsonify
from flask_jwt_extended import (
    JWTManager, create_access_token, create_refresh_token,
    jwt_required, get_jwt_identity, get_jwt
)
from flask_jwt_extended.utils import decode_token
import pyotp
from flask_cors import CORS
from dotenv import load_dotenv

# PYTHONPATH FIX
CURRENT_FILE = Path(__file__).resolve()
API_DIR = CURRENT_FILE.parent
BACKEND_DIR = API_DIR.parent

sys.path.insert(0, str(BACKEND_DIR))
sys.path.insert(0, str(API_DIR))

try:
    from db import get_connection, get_engine, get_sqlite_path
except ImportError:
    from backend.db import get_connection, get_engine, get_sqlite_path

# Imports locales
try:
    from auth import AuthManager
except ImportError:
    from api.auth import AuthManager

try:
    from core.notas_credito_manager import NotasCreditoManager
except ImportError:
    from backend.core.notas_credito_manager import NotasCreditoManager

# Configuración
load_dotenv()

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

# App Flask
app = Flask(__name__)

# JWT Configuration
JWT_SECRET = os.getenv('JWT_SECRET_KEY', 'CHANGE-THIS-SECRET-KEY-IN-PRODUCTION')
app.config['JWT_SECRET_KEY'] = JWT_SECRET
ACCESS_TOKEN_HOURS = int(os.getenv('JWT_ACCESS_TOKEN_HOURS', '8'))
REFRESH_TOKEN_DAYS = int(os.getenv('JWT_REFRESH_TOKEN_DAYS', '45'))
app.config['JWT_ACCESS_TOKEN_EXPIRES'] = timedelta(hours=ACCESS_TOKEN_HOURS)
app.config['JWT_REFRESH_TOKEN_EXPIRES'] = timedelta(days=REFRESH_TOKEN_DAYS)

jwt = JWTManager(app)

cors_origins_env = os.getenv('CORS_ORIGINS', '').strip()
if cors_origins_env:
    cors_origins = [origin.strip() for origin in cors_origins_env.split(',') if origin.strip()]
else:
    cors_origins = ["http://localhost:5173", "http://127.0.0.1:5173"]

CORS(
    app,
    resources={r"/api/*": {"origins": cors_origins}},
    supports_credentials=False,
    allow_headers=["Content-Type", "Authorization"],
    methods=["GET", "POST", "PUT", "DELETE", "OPTIONS"]
)

# Rate Limiter
try:
    from flask_limiter import Limiter
    from flask_limiter.util import get_remote_address
    limiter = Limiter(
        get_remote_address,
        app=app,
        default_limits=["200 per day", "50 per hour"],
        storage_uri="memory://"
    )
except Exception:
    class DummyLimiter:
        def limit(self, *args, **kwargs):
            def decorator(f):
                return f
            return decorator
    limiter = DummyLimiter()

auth_manager = AuthManager()

PROJECT_ROOT = BACKEND_DIR.parent
DB_PATH = Path(get_sqlite_path(str(PROJECT_ROOT / 'data' / 'notas_credito.db')))
notas_manager = NotasCreditoManager()

_cache_store = {}

def _get_cache(key: str):
    item = _cache_store.get(key)
    if not item:
        return None
    if item['expires'] < time.time():
        _cache_store.pop(key, None)
        return None
    return item['value']

def _set_cache(key: str, value, ttl_seconds: int):
    _cache_store[key] = {
        'value': value,
        'expires': time.time() + ttl_seconds
    }

def registrar_log(entidad: str, accion: str, entidad_id: str, usuario: str, payload: dict):
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute('''
            INSERT INTO audit_logs (entidad, accion, entidad_id, usuario, payload)
            VALUES (?, ?, ?, ?, ?)
        ''', (entidad, accion, entidad_id, usuario, json.dumps(payload, ensure_ascii=False)))
        conn.commit()
        conn.close()
    except Exception:
        pass

def _hash_nit(nit: str) -> str:
    return nit or ''

def _decrypt_value(value: str) -> str:
    return value or ''

def get_db_connection():
    return get_connection()

def _require_write_role():
    claims = get_jwt()
    if claims.get('rol') not in ('admin', 'editor'):
        return None, jsonify({"error": "No tiene permisos para modificar información"}), 403
    return claims, None, None


# JWT ERROR HANDLERS
@jwt.expired_token_loader
def expired_token_callback(jwt_header, jwt_payload):
    jti = jwt_payload.get('jti')
    token_type = jwt_payload.get('type')
    if jti and token_type == 'refresh':
        auth_manager.invalidar_sesion(jti)
    return jsonify({
        "error": "Token expirado",
        "message": "Tu sesión ha expirado. Por favor, inicia sesión nuevamente."
    }), 401


@jwt.invalid_token_loader
def invalid_token_callback(error):
    return jsonify({
        "error": "Token inválido",
        "message": "El token proporcionado no es válido."
    }), 401


@jwt.unauthorized_loader
def missing_token_callback(error):
    return jsonify({
        "error": "Token no proporcionado",
        "message": "Se requiere autenticación para acceder a este recurso."
    }), 401


@jwt.token_in_blocklist_loader
def check_if_token_revoked(jwt_header, jwt_payload):
    jti = jwt_payload.get('jti')
    if not jti:
        return True
    return not auth_manager.verificar_sesion_activa(jti)


# ENDPOINTS DE AUTENTICACIÓN
@app.route('/api/auth/login', methods=['POST'])
@limiter.limit("5 per minute")
def login():
    """Autenticación de usuario"""
    data = request.get_json()

    if not data or 'username' not in data or 'password' not in data:
        return jsonify({"error": "Username y password requeridos"}), 400

    username = data['username']
    password = data['password']
    otp = data.get('otp')
    ip_address = request.remote_addr or '0.0.0.0'

    autenticado, usuario, mensaje = auth_manager.autenticar(username, password, ip_address)

    if not autenticado:
        return jsonify({"error": mensaje}), 401

    twofa = auth_manager.obtener_2fa(usuario['id'])
    if twofa and twofa.get('habilitado'):
        if not otp:
            return jsonify({"error": "Se requiere OTP", "requires_2fa": True}), 206
        totp = pyotp.TOTP(twofa['secreto'])
        if not totp.verify(str(otp).strip()):
            return jsonify({"error": "OTP inválido", "requires_2fa": True}), 401

    user_id_str = str(usuario['id'])

    access_token = create_access_token(
        identity=user_id_str,
        additional_claims={
            'username': usuario['username'],
            'rol': usuario['rol'],
            'user_id': usuario['id']
        }
    )

    refresh_token = create_refresh_token(
        identity=user_id_str,
        additional_claims={'username': usuario['username']}
    )

    try:
        decoded_access = decode_token(access_token)
        decoded_refresh = decode_token(refresh_token)
        auth_manager.registrar_sesion(
            usuario['id'],
            decoded_access.get('jti'),
            decoded_refresh.get('jti'),
            ip_address,
            request.headers.get('User-Agent', ''),
            expires_in=REFRESH_TOKEN_DAYS * 24 * 3600
        )
    except Exception:
        pass

    return jsonify({
        "access_token": access_token,
        "refresh_token": refresh_token,
        "user": {
            "id": usuario['id'],
            "username": usuario['username'],
            "email": usuario['email'],
            "rol": usuario['rol']
        },
        "requires_2fa": False,
        "expires_in_seconds": ACCESS_TOKEN_HOURS * 3600
    }), 200


@app.route('/api/auth/refresh', methods=['POST'])
@jwt_required(refresh=True)
def refresh():
    """Obtener nuevo access token"""
    identity = str(get_jwt_identity())

    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute('SELECT id, username, rol FROM usuarios WHERE id = ? AND activo = 1', (identity,))
    row = cursor.fetchone()
    conn.close()

    if not row:
        return jsonify({"error": "Usuario no encontrado o inactivo"}), 401

    access_token = create_access_token(
        identity=identity,
        additional_claims={
            'username': row['username'],
            'rol': row['rol'],
            'user_id': row['id']
        }
    )

    return jsonify({"access_token": access_token}), 200


@app.route('/api/auth/logout', methods=['POST'])
@jwt_required()
def logout():
    """Cerrar sesión"""
    claims = get_jwt()
    jti = claims.get('jti')
    user_id = claims.get('user_id')
    if jti:
        auth_manager.invalidar_sesion(jti)
    if user_id:
        auth_manager.invalidar_sesiones_usuario(int(user_id))
    return jsonify({"mensaje": "Sesión cerrada exitosamente"}), 200


@app.route('/api/auth/change-password', methods=['POST'])
@jwt_required()
def change_password():
    """Cambiar contraseña del usuario actual"""
    data = request.get_json()

    if not data or 'nueva_contraseña' not in data:
        return jsonify({"error": "Nueva contraseña requerida"}), 400

    claims = get_jwt()
    username = claims.get('username')

    if auth_manager.cambiar_contraseña(username, data['nueva_contraseña']):
        registrar_log('usuarios', 'cambiar_contraseña', username, username, {})
        return jsonify({"mensaje": "Contraseña cambiada exitosamente"}), 200
    return jsonify({"error": "Error al cambiar contraseña"}), 500


@app.route('/api/auth/register', methods=['POST'])
@jwt_required()
def register_user():
    """Crear nuevo usuario - Solo admins"""
    claims = get_jwt()
    if claims.get('rol') != 'admin':
        return jsonify({"error": "No tiene permisos para crear usuarios"}), 403

    data = request.get_json()
    if not data or 'username' not in data or 'password' not in data:
        return jsonify({"error": "Username y password son requeridos"}), 400

    username = data['username']
    password = data['password']
    email = data.get('email')
    rol = data.get('rol', 'viewer')

    if rol not in ['admin', 'editor', 'viewer']:
        return jsonify({"error": "Rol inválido"}), 400

    if len(password) < 6:
        return jsonify({"error": "La contraseña debe tener al menos 6 caracteres"}), 400

    if auth_manager.crear_usuario(username, password, email, rol):
        registrar_log('usuarios', 'crear', username, claims.get('username'), {
            'username': username,
            'email': email,
            'rol': rol
        })
        return jsonify({
            "mensaje": "Usuario creado exitosamente",
            "usuario": {"username": username, "email": email, "rol": rol}
        }), 201
    return jsonify({"error": "Error al crear usuario. Puede que ya exista."}), 400


@app.route('/api/auth/2fa/setup', methods=['POST'])
@jwt_required()
def setup_2fa():
    try:
        claims = get_jwt()
        user_id = claims.get('user_id')
        username = claims.get('username')
        if not user_id:
            return jsonify({"error": "Usuario inválido"}), 400

        secret = pyotp.random_base32()
        auth_manager.guardar_2fa(int(user_id), secret, False)
        uri = pyotp.totp.TOTP(secret).provisioning_uri(name=username, issuer_name='CIPA')
        registrar_log('2fa', 'setup', str(user_id), username, {})
        return jsonify({"secret": secret, "otpauth_uri": uri}), 200
    except Exception as e:
        logger.error(f"Error en setup_2fa: {e}")
        return jsonify({"error": "Error al configurar 2FA"}), 500


@app.route('/api/auth/2fa/enable', methods=['POST'])
@jwt_required()
def enable_2fa():
    try:
        data = request.get_json() or {}
        otp = data.get('otp')
        claims = get_jwt()
        user_id = claims.get('user_id')
        username = claims.get('username')
        if not otp or not user_id:
            return jsonify({"error": "OTP requerido"}), 400
        record = auth_manager.obtener_2fa(int(user_id))
        if not record:
            return jsonify({"error": "2FA no configurado"}), 400
        totp = pyotp.TOTP(record['secreto'])
        if not totp.verify(str(otp).strip()):
            return jsonify({"error": "OTP inválido"}), 401
        auth_manager.actualizar_2fa(int(user_id), True)
        registrar_log('2fa', 'enable', str(user_id), username, {})
        return jsonify({"mensaje": "2FA habilitado"}), 200
    except Exception as e:
        logger.error(f"Error en enable_2fa: {e}")
        return jsonify({"error": "Error al habilitar 2FA"}), 500


@app.route('/api/auth/2fa/disable', methods=['POST'])
@jwt_required()
def disable_2fa():
    try:
        claims = get_jwt()
        user_id = claims.get('user_id')
        username = claims.get('username')
        if not user_id:
            return jsonify({"error": "Usuario inválido"}), 400
        auth_manager.actualizar_2fa(int(user_id), False)
        registrar_log('2fa', 'disable', str(user_id), username, {})
        return jsonify({"mensaje": "2FA deshabilitado"}), 200
    except Exception as e:
        logger.error(f"Error en disable_2fa: {e}")
        return jsonify({"error": "Error al deshabilitar 2FA"}), 500


@app.route('/api/auth/2fa/status', methods=['GET'])
@jwt_required()
def status_2fa():
    try:
        claims = get_jwt()
        user_id = claims.get('user_id')
        if not user_id:
            return jsonify({"error": "Usuario inválido"}), 400
        record = auth_manager.obtener_2fa(int(user_id))
        return jsonify({
            "configurado": record is not None,
            "habilitado": bool(record.get('habilitado')) if record else False
        }), 200
    except Exception as e:
        logger.error(f"Error en status_2fa: {e}")
        return jsonify({"error": "Error al consultar 2FA"}), 500


@app.route('/api/auth/users', methods=['GET'])
@jwt_required()
def list_users():
    """Listar usuarios - Solo admins"""
    claims = get_jwt()
    if claims.get('rol') != 'admin':
        return jsonify({"error": "No tiene permisos para ver usuarios"}), 403

    conn = get_db_connection()
    cursor = conn.cursor()

    cursor.execute('''
        SELECT id, username, email, rol, activo, ultimo_acceso, fecha_creacion
        FROM usuarios ORDER BY fecha_creacion DESC
    ''')

    usuarios = []
    for row in cursor.fetchall():
        usuarios.append({
            'id': row[0],
            'username': row[1],
            'email': row[2],
            'rol': row[3],
            'activo': bool(row[4]),
            'ultimo_acceso': row[5],
            'fecha_creacion': row[6]
        })

    conn.close()
    return jsonify({"total": len(usuarios), "usuarios": usuarios}), 200


# ENDPOINTS DE FACTURAS
@app.route('/api/facturas', methods=['GET'])
@jwt_required()
def listar_facturas():
    try:
        nit_cliente = request.args.get('nit_cliente')
        fecha_desde = request.args.get('fecha_desde')
        fecha_hasta = request.args.get('fecha_hasta')
        con_nota = request.args.get('con_nota')
        codigo_factura = request.args.get('codigo_factura')
        numero_factura = request.args.get('numero_factura')
        nombre_cliente = request.args.get('nombre_cliente')
        estado = request.args.get('estado')
        registrable = request.args.get('registrable')
        search = request.args.get('search')
        orden = request.args.get('orden', 'fecha_factura')
        direccion = request.args.get('direccion', 'DESC')
        limite = int(request.args.get('limite', 100))
        offset = int(request.args.get('offset', 0))

        query = "SELECT * FROM facturas WHERE 1=1"
        params = []

        if nit_cliente:
            query += " AND nit_hash = ?"
            params.append(_hash_nit(nit_cliente))

        if fecha_desde:
            query += " AND fecha_factura >= ?"
            params.append(fecha_desde)

        if fecha_hasta:
            query += " AND fecha_factura <= ?"
            params.append(fecha_hasta)

        if codigo_factura:
            query += " AND codigo_factura = ?"
            params.append(codigo_factura)

        if numero_factura:
            query += " AND numero_factura = ?"
            params.append(numero_factura)

        if nombre_cliente:
            query += " AND nombre_cliente_encrypted LIKE ?"
            params.append(f"%{nombre_cliente}%")

        if estado:
            query += " AND estado = ?"
            params.append(estado)

        if registrable is not None:
            if registrable.lower() == 'true' or registrable == '1':
                query += " AND registrable = 1"
            elif registrable.lower() == 'false' or registrable == '0':
                query += " AND registrable = 0"

        if con_nota is not None:
            if con_nota.lower() == 'true' or con_nota == '1':
                query += " AND EXISTS (SELECT 1 FROM aplicaciones_notas a WHERE a.numero_factura = facturas.numero_factura)"
            elif con_nota.lower() == 'false' or con_nota == '0':
                query += " AND NOT EXISTS (SELECT 1 FROM aplicaciones_notas a WHERE a.numero_factura = facturas.numero_factura)"

        if search:
            like = f"%{search}%"
            query += " AND (numero_factura LIKE ? OR nit_encrypted LIKE ? OR nombre_cliente_encrypted LIKE ? OR codigo_factura LIKE ?)"
            params.extend([like, like, like, like])

        orden_map = {
            'fecha_factura': 'fecha_factura',
            'numero_factura': 'numero_factura',
            'valor_total': 'valor_total',
            'nombre_cliente': 'nombre_cliente_encrypted',
            'nit_cliente': 'nit_encrypted',
            'codigo_factura': 'codigo_factura'
        }
        orden_col = orden_map.get(orden, 'fecha_factura')
        direccion_sql = 'ASC' if str(direccion).upper() == 'ASC' else 'DESC'
        query += f" ORDER BY {orden_col} {direccion_sql} LIMIT ? OFFSET ?"
        params.extend([limite, offset])

        conn = get_db_connection()
        cursor = conn.cursor()

        cursor.execute(query, params)
        facturas = []
        for row in cursor.fetchall():
            item = dict(row)
            item['nit_cliente'] = _decrypt_value(item.get('nit_encrypted'))
            item['nombre_cliente'] = _decrypt_value(item.get('nombre_cliente_encrypted'))
            cursor.execute('''
                SELECT numero_nota, valor_aplicado, fecha_aplicacion
                FROM aplicaciones_notas
                WHERE id_factura = ?
                ORDER BY fecha_aplicacion DESC
                LIMIT 1
            ''', (item['id'],))
            ultima_aplicacion = cursor.fetchone()
            if ultima_aplicacion:
                item['ultima_nota_aplicada'] = ultima_aplicacion['numero_nota']
                item['monto_ultima_aplicacion'] = float(ultima_aplicacion['valor_aplicado'] or 0)
                item['fecha_ultima_aplicacion'] = ultima_aplicacion['fecha_aplicacion']
            else:
                item['ultima_nota_aplicada'] = None
                item['monto_ultima_aplicacion'] = 0.0
                item['fecha_ultima_aplicacion'] = None

            cursor.execute('''
                SELECT COALESCE(SUM(valor_aplicado), 0) AS total_aplicado
                FROM aplicaciones_notas
                WHERE id_factura = ?
            ''', (item['id'],))
            total_aplicado_row = cursor.fetchone()
            item['monto_total_aplicado'] = float(total_aplicado_row['total_aplicado']) if total_aplicado_row else 0.0
            facturas.append(item)

        query_count = query.split('ORDER BY')[0].replace('SELECT *', 'SELECT COUNT(*)')
        cursor.execute(query_count, params[:-2])
        total = cursor.fetchone()[0]

        conn.close()

        return jsonify({
            "items": facturas,
            "total": total,
            "limite": limite,
            "offset": offset,
            "total_paginas": (total + limite - 1) // limite
        }), 200

    except Exception as e:
        logger.error(f"Error en listar_facturas: {e}")
        return jsonify({"error": "Error al obtener facturas"}), 500


@app.route('/api/facturas/<int:factura_id>', methods=['GET'])
@jwt_required()
def obtener_factura(factura_id):
    try:
        conn = get_db_connection()
        cursor = conn.cursor()

        cursor.execute('SELECT * FROM facturas WHERE id = ?', (factura_id,))
        factura = cursor.fetchone()

        if not factura:
            conn.close()
            return jsonify({"error": "Factura no encontrada"}), 404

        factura_dict = dict(factura)
        factura_dict['nit_cliente'] = _decrypt_value(factura_dict.get('nit_encrypted'))
        factura_dict['nombre_cliente'] = _decrypt_value(factura_dict.get('nombre_cliente_encrypted'))

        # Obtener aplicaciones de notas
        cursor.execute('''
            SELECT * FROM aplicaciones_notas
            WHERE numero_factura = ?
            ORDER BY fecha_aplicacion DESC
        ''', (factura['numero_factura'],))

        factura_dict['aplicaciones'] = [dict(row) for row in cursor.fetchall()]
        conn.close()

        return jsonify(factura_dict), 200

    except Exception as e:
        logger.error(f"Error en obtener_factura: {e}")
        return jsonify({"error": "Error al obtener factura"}), 500


@app.route('/api/facturas', methods=['POST'])
@jwt_required()
def crear_factura():
    try:
        claims, error_response, status_code = _require_write_role()
        if error_response:
            return error_response, status_code
        data = request.get_json() or {}
        usuario = claims.get('username')

        if notas_manager.registrar_factura(data, usuario):
            return jsonify({"mensaje": "Factura creada"}), 201
        return jsonify({"error": "No se pudo crear la factura"}), 400
    except Exception as e:
        logger.error(f"Error en crear_factura: {e}")
        return jsonify({"error": "Error al crear factura"}), 500


@app.route('/api/facturas/<int:factura_id>', methods=['PUT'])
@jwt_required()
def actualizar_factura(factura_id):
    try:
        claims, error_response, status_code = _require_write_role()
        if error_response:
            return error_response, status_code
        data = request.get_json() or {}
        usuario = claims.get('username')

        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute('SELECT * FROM facturas WHERE id = ?', (factura_id,))
        factura = cursor.fetchone()
        if not factura:
            conn.close()
            return jsonify({"error": "Factura no encontrada"}), 404

        numero_factura = str(data.get('numero_factura', factura['numero_factura'])).strip()
        codigo_factura = str(data.get('codigo_factura', factura['codigo_factura'])).strip()
        codigo_producto = str(data.get('codigo_producto', factura['codigo_producto'])).strip()
        nombre_producto = str(data.get('nombre_producto', factura['nombre_producto'])).strip()
        nit_cliente = str(data.get('nit_cliente', _decrypt_value(factura['nit_encrypted']))).strip()
        nombre_cliente = str(data.get('nombre_cliente', _decrypt_value(factura['nombre_cliente_encrypted']))).strip()
        cantidad_original = float(data.get('cantidad_original', factura['cantidad_original']) or 0.0)
        valor_total = float(data.get('valor_total', factura['valor_total']) or 0.0)
        fecha_factura = data.get('fecha_factura', factura['fecha_factura'])

        if not numero_factura or not codigo_producto or not nit_cliente:
            conn.close()
            return jsonify({"error": "Datos incompletos"}), 400

        notas_manager._validate_plain_text(nit_cliente, 'nit_cliente', numero_factura)
        notas_manager._validate_plain_text(nombre_cliente, 'nombre_cliente', numero_factura)
        notas_manager._validate_nit(nit_cliente, numero_factura)

        nit_hash = _hash_nit(nit_cliente)
        nit_enc = nit_cliente
        nombre_enc = nombre_cliente

        codigo_factura_norm = codigo_factura.lower()
        registrable = 1
        total_repeticiones = 1
        suma_total_repeticiones = valor_total

        if codigo_factura_norm == 'abc123':
            cursor.execute('''
                SELECT COUNT(*), COALESCE(SUM(valor_total), 0)
                FROM facturas WHERE codigo_factura = ? AND id != ?
            ''', (codigo_factura, factura_id))
            count, suma = cursor.fetchone()
            if count >= 5:
                conn.close()
                return jsonify({"error": "Límite de repeticiones alcanzado"}), 400
            total_repeticiones = count + 1
            suma_total_repeticiones = float(suma or 0) + valor_total
            registrable = 1 if suma_total_repeticiones > 524000 else 0
        else:
            if valor_total < 524000:
                conn.close()
                return jsonify({"error": "Monto mínimo no alcanzado"}), 400

        cursor.execute('''
            UPDATE facturas
            SET numero_factura = ?, codigo_factura = ?, codigo_producto = ?, nombre_producto = ?,
                nit_encrypted = ?, nit_hash = ?, nombre_cliente_encrypted = ?,
                cantidad_original = ?, valor_total = ?, cantidad_restante = ?, valor_restante = ?,
                registrable = ?, total_repeticiones = ?, suma_total_repeticiones = ?, fecha_factura = ?
            WHERE id = ?
        ''', (
            numero_factura, codigo_factura, codigo_producto, nombre_producto,
            nit_enc, nit_hash, nombre_enc,
            cantidad_original, valor_total, cantidad_original, valor_total,
            registrable, total_repeticiones, suma_total_repeticiones, fecha_factura,
            factura_id
        ))

        if codigo_factura_norm == 'abc123':
            cursor.execute('''
                UPDATE facturas
                SET total_repeticiones = ?, suma_total_repeticiones = ?, registrable = ?
                WHERE codigo_factura = ?
            ''', (total_repeticiones, suma_total_repeticiones, registrable, codigo_factura))

        cursor.execute('''
            INSERT INTO audit_logs (entidad, accion, entidad_id, usuario, payload)
            VALUES (?, ?, ?, ?, ?)
        ''', ('factura', 'actualizar', str(factura_id), usuario, json.dumps(data, ensure_ascii=False)))

        conn.commit()
        conn.close()
        return jsonify({"mensaje": "Factura actualizada"}), 200

    except Exception as e:
        logger.error(f"Error en actualizar_factura: {e}")
        return jsonify({"error": "Error al actualizar factura"}), 500


@app.route('/api/facturas/<int:factura_id>', methods=['DELETE'])
@jwt_required()
def eliminar_factura(factura_id):
    try:
        claims, error_response, status_code = _require_write_role()
        if error_response:
            return error_response, status_code
        usuario = claims.get('username')

        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute('SELECT numero_factura FROM facturas WHERE id = ?', (factura_id,))
        row = cursor.fetchone()
        if not row:
            conn.close()
            return jsonify({"error": "Factura no encontrada"}), 404

        cursor.execute('DELETE FROM facturas WHERE id = ?', (factura_id,))
        cursor.execute('''
            INSERT INTO audit_logs (entidad, accion, entidad_id, usuario, payload)
            VALUES (?, ?, ?, ?, ?)
        ''', ('factura', 'eliminar', str(factura_id), usuario, json.dumps({'numero_factura': row['numero_factura']}, ensure_ascii=False)))

        conn.commit()
        conn.close()
        return jsonify({"mensaje": "Factura eliminada"}), 200
    except Exception as e:
        logger.error(f"Error en eliminar_factura: {e}")
        return jsonify({"error": "Error al eliminar factura"}), 500


@app.route('/api/facturas/estadisticas', methods=['GET'])
@jwt_required()
def estadisticas_facturas():
    """Estadísticas de facturas"""
    try:
        conn = get_db_connection()
        cursor = conn.cursor()

        stats = {}

        cursor.execute('SELECT COUNT(*), SUM(valor_total) FROM facturas')
        row = cursor.fetchone()
        stats['facturas_validas'] = row[0] or 0
        stats['valor_total_facturado'] = row[1] or 0

        cursor.execute('SELECT COUNT(*) FROM facturas WHERE registrable = 1')
        stats['facturas_registrables'] = cursor.fetchone()[0] or 0

        cursor.execute('SELECT COUNT(*) FROM facturas WHERE registrable = 0')
        stats['facturas_no_registrables'] = cursor.fetchone()[0] or 0

        cursor.execute('SELECT COUNT(*), SUM(valor_aplicado) FROM aplicaciones_notas')
        row = cursor.fetchone()
        stats['aplicaciones_total'] = row[0] or 0
        stats['total_aplicado'] = row[1] or 0

        cursor.execute('SELECT COUNT(*) FROM facturas_rechazadas')
        stats['facturas_rechazadas'] = cursor.fetchone()[0]

        cursor.execute('SELECT SUM(valor_total) FROM facturas_rechazadas')
        stats['valor_rechazado'] = cursor.fetchone()[0] or 0

        conn.close()
        return jsonify(stats), 200

    except Exception as e:
        logger.error(f"Error en estadisticas_facturas: {e}")
        return jsonify({"error": "Error al obtener estadísticas"}), 500


@app.route('/api/facturas/transacciones', methods=['GET'])
@jwt_required()
def listar_transacciones():
    try:
        limite = int(request.args.get('limite', 50))
        offset = int(request.args.get('offset', 0))

        conn = get_db_connection()
        cursor = conn.cursor()

        cursor.execute('''
            SELECT a.id, a.numero_nota, a.numero_factura, a.cantidad_aplicada, a.valor_aplicado, a.fecha_aplicacion,
                   n.nit_encrypted
            FROM aplicaciones_notas a
            LEFT JOIN notas_credito n ON n.id = a.id_nota
            ORDER BY a.fecha_aplicacion DESC
            LIMIT ? OFFSET ?
        ''', (limite, offset))
        items = []
        for row in cursor.fetchall():
            item = dict(row)
            item['nit_cliente'] = _decrypt_value(item.get('nit_encrypted'))
            items.append(item)

        cursor.execute('SELECT COUNT(*) FROM aplicaciones_notas')
        total = cursor.fetchone()[0]
        conn.close()

        return jsonify({
            "items": items,
            "total": total,
            "limite": limite,
            "offset": offset
        }), 200
    except Exception as e:
        logger.error(f"Error en listar_transacciones: {e}")
        return jsonify({"error": "Error al obtener transacciones"}), 500


@app.route('/api/facturas/transado-mensual', methods=['GET'])
@jwt_required()
def transado_mensual_facturas():
    try:
        limite = int(request.args.get('limite', 12))
        if limite <= 0:
            limite = 12
        conn = get_db_connection()
        cursor = conn.cursor()

        if conn.engine == 'mysql':
            cursor.execute('''
                SELECT DATE_FORMAT(fecha_factura, '%Y-%m') AS periodo,
                       SUM(valor_total) AS valor_total,
                       COUNT(*) AS total_facturas
                FROM facturas
                WHERE registrable = 1 AND (estado = 'ACTIVA' OR estado IS NULL)
                GROUP BY DATE_FORMAT(fecha_factura, '%Y-%m')
                ORDER BY periodo DESC
                LIMIT ?
            ''', (limite,))
        else:
            cursor.execute('''
                SELECT SUBSTR(fecha_factura, 1, 7) AS periodo,
                       SUM(valor_total) AS valor_total,
                       COUNT(*) AS total_facturas
                FROM facturas
                WHERE registrable = 1 AND (estado = 'ACTIVA' OR estado IS NULL)
                GROUP BY SUBSTR(fecha_factura, 1, 7)
                ORDER BY periodo DESC
                LIMIT ?
            ''', (limite,))

        rows = cursor.fetchall()
        conn.close()

        items = []
        for row in reversed(rows):
            periodo = str(row['periodo']) if row.get('periodo') else ''
            items.append({
                "periodo": periodo,
                "valor_total": float(row['valor_total'] or 0),
                "total_facturas": int(row['total_facturas'] or 0)
            })

        return jsonify({
            "items": items,
            "limite": limite
        }), 200
    except Exception as e:
        logger.error(f"Error en transado_mensual_facturas: {e}")
        return jsonify({"error": "Error al obtener transado mensual"}), 500


@app.route('/api/notas/aplicado-mensual', methods=['GET'])
@jwt_required()
def aplicado_mensual_notas():
    try:
        limite = int(request.args.get('limite', 12))
        if limite <= 0:
            limite = 12
        conn = get_db_connection()
        cursor = conn.cursor()

        if conn.engine == 'mysql':
            cursor.execute('''
                SELECT DATE_FORMAT(fecha_aplicacion, '%Y-%m') AS periodo,
                       SUM(valor_aplicado) AS valor_aplicado,
                       COUNT(*) AS total_aplicaciones
                FROM aplicaciones_notas
                GROUP BY DATE_FORMAT(fecha_aplicacion, '%Y-%m')
                ORDER BY periodo DESC
                LIMIT ?
            ''', (limite,))
        else:
            cursor.execute('''
                SELECT SUBSTR(fecha_aplicacion, 1, 7) AS periodo,
                       SUM(valor_aplicado) AS valor_aplicado,
                       COUNT(*) AS total_aplicaciones
                FROM aplicaciones_notas
                GROUP BY SUBSTR(fecha_aplicacion, 1, 7)
                ORDER BY periodo DESC
                LIMIT ?
            ''', (limite,))

        rows = cursor.fetchall()
        conn.close()

        items = []
        for row in reversed(rows):
            periodo = str(row['periodo']) if row.get('periodo') else ''
            items.append({
                "periodo": periodo,
                "valor_aplicado": float(row['valor_aplicado'] or 0),
                "total_aplicaciones": int(row['total_aplicaciones'] or 0)
            })

        return jsonify({
            "items": items,
            "limite": limite
        }), 200
    except Exception as e:
        logger.error(f"Error en aplicado_mensual_notas: {e}")
        return jsonify({"error": "Error al obtener aplicado mensual de notas"}), 500


@app.route('/api/facturas/rechazadas', methods=['GET'])
@jwt_required()
def listar_facturas_rechazadas():
    """Listar facturas rechazadas"""
    try:
        limite = int(request.args.get('limite', 50))
        offset = int(request.args.get('offset', 0))
        fecha_desde = request.args.get('fecha_desde')
        fecha_hasta = request.args.get('fecha_hasta')

        query = "SELECT * FROM facturas_rechazadas WHERE 1=1"
        params = []

        if fecha_desde:
            query += " AND fecha_factura >= ?"
            params.append(fecha_desde)

        if fecha_hasta:
            query += " AND fecha_factura <= ?"
            params.append(fecha_hasta)

        query += " ORDER BY fecha_registro DESC LIMIT ? OFFSET ?"
        params.extend([limite, offset])

        conn = get_db_connection()
        cursor = conn.cursor()

        cursor.execute(query, params)
        rechazadas = []
        for row in cursor.fetchall():
            item = dict(row)
            item['nit_cliente'] = _decrypt_value(item.get('nit_encrypted'))
            item['nombre_cliente'] = _decrypt_value(item.get('nombre_cliente_encrypted'))
            rechazadas.append(item)

        query_count = query.split('ORDER BY')[0].replace('SELECT *', 'SELECT COUNT(*)')
        cursor.execute(query_count, params[:-2])
        total = cursor.fetchone()[0]

        conn.close()

        return jsonify({
            "items": rechazadas,
            "total": total,
            "limite": limite,
            "offset": offset
        }), 200

    except Exception as e:
        logger.error(f"Error en listar_facturas_rechazadas: {e}")
        return jsonify({"error": "Error al obtener facturas rechazadas"}), 500


# ENDPOINTS DE NOTAS CRÉDITO
@app.route('/api/notas', methods=['GET'])
@jwt_required()
def listar_notas():
    """Listar notas de crédito"""
    try:
        estado = request.args.get('estado')
        nit_cliente = request.args.get('nit_cliente')
        fecha_desde = request.args.get('fecha_desde')
        fecha_hasta = request.args.get('fecha_hasta')
        limite = int(request.args.get('limite', 100))
        offset = int(request.args.get('offset', 0))

        query = "SELECT * FROM notas_credito WHERE 1=1"
        params = []

        if estado:
            query += " AND estado = ?"
            params.append(estado)

        if nit_cliente:
            query += " AND nit_hash = ?"
            params.append(_hash_nit(nit_cliente))

        if fecha_desde:
            query += " AND fecha_nota >= ?"
            params.append(fecha_desde)

        if fecha_hasta:
            query += " AND fecha_nota <= ?"
            params.append(fecha_hasta)

        query += " ORDER BY fecha_nota DESC LIMIT ? OFFSET ?"
        params.extend([limite, offset])

        conn = get_db_connection()
        cursor = conn.cursor()

        cursor.execute(query, params)
        notas = []
        for row in cursor.fetchall():
            item = dict(row)
            item['nit_cliente'] = _decrypt_value(item.get('nit_encrypted'))
            item['nombre_cliente'] = _decrypt_value(item.get('nombre_cliente_encrypted'))
            cursor.execute('''
                SELECT numero_factura, valor_aplicado, fecha_aplicacion
                FROM aplicaciones_notas
                WHERE id_nota = ?
                ORDER BY fecha_aplicacion DESC
                LIMIT 1
            ''', (item['id'],))
            ultima_aplicacion = cursor.fetchone()
            if ultima_aplicacion:
                item['factura_aplicada'] = ultima_aplicacion['numero_factura']
                item['monto_aplicado'] = float(ultima_aplicacion['valor_aplicado'] or 0)
                item['fecha_ultima_aplicacion'] = ultima_aplicacion['fecha_aplicacion']
            else:
                item['factura_aplicada'] = None
                item['monto_aplicado'] = 0.0
                item['fecha_ultima_aplicacion'] = None

            item['tipo_aplicacion'] = 'COMPLETA' if float(item.get('saldo_pendiente') or 0) <= 0 and item.get('estado') == 'APLICADA' else 'PARCIAL'
            notas.append(item)

        query_count = query.split('ORDER BY')[0].replace('SELECT *', 'SELECT COUNT(*)')
        cursor.execute(query_count, params[:-2])
        total = cursor.fetchone()[0]

        conn.close()

        return jsonify({
            "items": notas,
            "total": total,
            "limite": limite,
            "offset": offset,
            "total_paginas": (total + limite - 1) // limite
        }), 200

    except Exception as e:
        logger.error(f"Error en listar_notas: {e}")
        return jsonify({"error": "Error al obtener notas"}), 500


@app.route('/api/notas/<int:nota_id>', methods=['GET'])
@jwt_required()
def obtener_nota(nota_id):
    """Obtener detalles de una nota"""
    try:
        conn = get_db_connection()
        cursor = conn.cursor()

        cursor.execute('SELECT * FROM notas_credito WHERE id = ?', (nota_id,))
        nota = cursor.fetchone()

        if not nota:
            conn.close()
            return jsonify({"error": "Nota no encontrada"}), 404

        nota_dict = dict(nota)
        nota_dict['nit_cliente'] = _decrypt_value(nota_dict.get('nit_encrypted'))
        nota_dict['nombre_cliente'] = _decrypt_value(nota_dict.get('nombre_cliente_encrypted'))

        # Obtener historial de aplicaciones
        cursor.execute('''
            SELECT * FROM aplicaciones_notas
            WHERE id_nota = ?
            ORDER BY fecha_aplicacion DESC
        ''', (nota_id,))

        nota_dict['aplicaciones'] = [dict(row) for row in cursor.fetchall()]

        cursor.execute('''
            SELECT * FROM log_motivos_no_aplicacion
            WHERE id_nota = ?
            ORDER BY fecha_registro DESC
        ''', (nota_id,))
        nota_dict['motivos_no_aplicacion'] = [dict(row) for row in cursor.fetchall()]
        conn.close()

        return jsonify(nota_dict), 200

    except Exception as e:
        logger.error(f"Error en obtener_nota: {e}")
        return jsonify({"error": "Error al obtener nota"}), 500


@app.route('/api/notas', methods=['POST'])
@jwt_required()
def crear_nota():
    try:
        claims, error_response, status_code = _require_write_role()
        if error_response:
            return error_response, status_code
        data = request.get_json() or {}
        usuario = claims.get('username')

        if notas_manager.registrar_nota_credito(data, usuario):
            return jsonify({"mensaje": "Nota creada"}), 201
        return jsonify({"error": "No se pudo crear la nota"}), 400
    except Exception as e:
        logger.error(f"Error en crear_nota: {e}")
        return jsonify({"error": "Error al crear nota"}), 500


@app.route('/api/notas/<int:nota_id>', methods=['PUT'])
@jwt_required()
def actualizar_nota(nota_id):
    try:
        claims, error_response, status_code = _require_write_role()
        if error_response:
            return error_response, status_code
        data = request.get_json() or {}
        usuario = claims.get('username')

        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute('SELECT * FROM notas_credito WHERE id = ?', (nota_id,))
        nota = cursor.fetchone()
        if not nota:
            conn.close()
            return jsonify({"error": "Nota no encontrada"}), 404

        numero_nota = str(data.get('numero_nota', nota['numero_nota'])).strip()
        codigo_producto = str(data.get('codigo_producto', nota['codigo_producto'])).strip()
        nombre_producto = str(data.get('nombre_producto', nota['nombre_producto'])).strip()
        nit_cliente = str(data.get('nit_cliente', _decrypt_value(nota['nit_encrypted']))).strip()
        nombre_cliente = str(data.get('nombre_cliente', _decrypt_value(nota['nombre_cliente_encrypted']))).strip()
        valor_total = float(data.get('valor_total', nota['valor_total']) or 0.0)
        cantidad = float(data.get('cantidad', nota['cantidad']) or 0.0)
        estado = str(data.get('estado', nota['estado'])).strip()
        es_agente = 1 if str(data.get('es_agente', nota['es_agente'])).lower() in ('1', 'true', 'si', 'sí') else 0

        notas_manager._validate_plain_text(nit_cliente, 'nit_cliente', numero_nota)
        notas_manager._validate_plain_text(nombre_cliente, 'nombre_cliente', numero_nota)
        notas_manager._validate_nit(nit_cliente, numero_nota)

        nit_hash = _hash_nit(nit_cliente)
        nit_enc = nit_cliente
        nombre_enc = nombre_cliente

        cursor.execute('''
            UPDATE notas_credito
            SET numero_nota = ?, codigo_producto = ?, nombre_producto = ?,
                nit_encrypted = ?, nit_hash = ?, nombre_cliente_encrypted = ?,
                valor_total = ?, cantidad = ?, saldo_pendiente = ?, cantidad_pendiente = ?,
                estado = ?, es_agente = ?
            WHERE id = ?
        ''', (
            numero_nota, codigo_producto, nombre_producto,
            nit_enc, nit_hash, nombre_enc,
            valor_total, cantidad, valor_total, cantidad,
            estado, es_agente, nota_id
        ))

        cursor.execute('''
            INSERT INTO audit_logs (entidad, accion, entidad_id, usuario, payload)
            VALUES (?, ?, ?, ?, ?)
        ''', ('nota_credito', 'actualizar', str(nota_id), usuario, json.dumps(data, ensure_ascii=False)))

        conn.commit()
        conn.close()
        return jsonify({"mensaje": "Nota actualizada"}), 200
    except Exception as e:
        logger.error(f"Error en actualizar_nota: {e}")
        return jsonify({"error": "Error al actualizar nota"}), 500


@app.route('/api/notas/<int:nota_id>', methods=['DELETE'])
@jwt_required()
def eliminar_nota(nota_id):
    try:
        claims, error_response, status_code = _require_write_role()
        if error_response:
            return error_response, status_code
        usuario = claims.get('username')

        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute('SELECT numero_nota FROM notas_credito WHERE id = ?', (nota_id,))
        row = cursor.fetchone()
        if not row:
            conn.close()
            return jsonify({"error": "Nota no encontrada"}), 404

        cursor.execute('DELETE FROM notas_credito WHERE id = ?', (nota_id,))
        cursor.execute('''
            INSERT INTO audit_logs (entidad, accion, entidad_id, usuario, payload)
            VALUES (?, ?, ?, ?, ?)
        ''', ('nota_credito', 'eliminar', str(nota_id), usuario, json.dumps({'numero_nota': row['numero_nota']}, ensure_ascii=False)))

        conn.commit()
        conn.close()
        return jsonify({"mensaje": "Nota eliminada"}), 200
    except Exception as e:
        logger.error(f"Error en eliminar_nota: {e}")
        return jsonify({"error": "Error al eliminar nota"}), 500


@app.route('/api/notas/aplicar', methods=['POST'])
@jwt_required()
def aplicar_nota():
    try:
        claims, error_response, status_code = _require_write_role()
        if error_response:
            return error_response, status_code
        data = request.get_json() or {}
        usuario = claims.get('username')

        nota_id = data.get('nota_id')
        numero_factura = data.get('numero_factura')
        codigo_producto = data.get('codigo_producto')
        indice_linea = data.get('indice_linea', 0)

        if not nota_id or not numero_factura or not codigo_producto:
            return jsonify({"error": "Datos incompletos"}), 400

        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute('SELECT * FROM notas_credito WHERE id = ?', (nota_id,))
        nota = cursor.fetchone()
        if not nota:
            conn.close()
            return jsonify({"error": "Nota no encontrada"}), 404

        cursor.execute('''
            SELECT * FROM facturas
            WHERE numero_factura = ? AND codigo_producto = ? AND indice_linea = ?
        ''', (numero_factura, codigo_producto, indice_linea))
        factura = cursor.fetchone()
        if not factura:
            conn.close()
            return jsonify({"error": "Factura no encontrada"}), 404

        nota_dict = dict(nota)
        nota_dict['nit_cliente'] = _decrypt_value(nota_dict.get('nit_encrypted'))
        nota_dict['nombre_cliente'] = _decrypt_value(nota_dict.get('nombre_cliente_encrypted'))

        factura_dict = dict(factura)
        factura_dict['nit_comprador'] = _decrypt_value(factura_dict.get('nit_encrypted'))
        factura_dict['nombre_comprador'] = _decrypt_value(factura_dict.get('nombre_cliente_encrypted'))
        factura_dict['codigo_producto_api'] = factura_dict.get('codigo_producto')
        factura_dict['cantidad_original'] = factura_dict.get('cantidad_restante')
        factura_dict['valor_total'] = factura_dict.get('valor_restante')

        conn.close()

        resultado = notas_manager.aplicar_nota_a_factura(nota_dict, factura_dict, usuario)
        if not resultado:
            return jsonify({"error": "No se pudo aplicar la nota"}), 400
        return jsonify(resultado), 200

    except Exception as e:
        logger.error(f"Error en aplicar_nota: {e}")
        return jsonify({"error": "Error al aplicar nota"}), 500


@app.route('/api/notas/no-aplicadas', methods=['GET'])
@jwt_required()
def listar_no_aplicadas():
    try:
        limite = int(request.args.get('limite', 100))
        offset = int(request.args.get('offset', 0))
        conn = get_db_connection()
        cursor = conn.cursor()

        cursor.execute('''
            SELECT * FROM log_motivos_no_aplicacion
            ORDER BY fecha_registro DESC LIMIT ? OFFSET ?
        ''', (limite, offset))
        items = [dict(row) for row in cursor.fetchall()]

        cursor.execute('SELECT COUNT(*) FROM log_motivos_no_aplicacion')
        total = cursor.fetchone()[0]

        conn.close()
        return jsonify({"items": items, "total": total, "limite": limite, "offset": offset}), 200
    except Exception as e:
        logger.error(f"Error en listar_no_aplicadas: {e}")
        return jsonify({"error": "Error al obtener no aplicadas"}), 500


@app.route('/api/notas/no-aplicadas', methods=['POST'])
@jwt_required()
def registrar_no_aplicada():
    try:
        claims, error_response, status_code = _require_write_role()
        if error_response:
            return error_response, status_code
        data = request.get_json() or {}
        usuario = claims.get('username')

        nota_id = data.get('nota_id')
        motivo = data.get('motivo')
        numero_factura = data.get('numero_factura')
        detalle = data.get('detalle')

        if not nota_id or not motivo:
            return jsonify({"error": "Datos incompletos"}), 400

        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute('SELECT * FROM notas_credito WHERE id = ?', (nota_id,))
        nota = cursor.fetchone()
        if not nota:
            conn.close()
            return jsonify({"error": "Nota no encontrada"}), 404

        cursor.execute('''
            INSERT INTO log_motivos_no_aplicacion
            (id_nota, numero_nota, numero_factura, motivo, detalle)
            VALUES (?, ?, ?, ?, ?)
        ''', (nota_id, nota['numero_nota'], numero_factura, motivo, detalle))

        cursor.execute('''
            UPDATE notas_credito SET estado = 'NO_APLICADA' WHERE id = ?
        ''', (nota_id,))

        cursor.execute('''
            INSERT INTO audit_logs (entidad, accion, entidad_id, usuario, payload)
            VALUES (?, ?, ?, ?, ?)
        ''', ('nota_credito', 'no_aplicada', str(nota_id), usuario, json.dumps({'motivo': motivo}, ensure_ascii=False)))

        conn.commit()
        conn.close()
        return jsonify({"mensaje": "Motivo registrado"}), 201
    except Exception as e:
        logger.error(f"Error en registrar_no_aplicada: {e}")
        return jsonify({"error": "Error al registrar motivo"}), 500


@app.route('/api/notas/estadisticas', methods=['GET'])
@jwt_required()
def estadisticas_notas():
    """Estadísticas de notas de crédito"""
    try:
        cached = _get_cache('notas_stats')
        if cached:
            return jsonify(cached), 200

        conn = get_db_connection()
        cursor = conn.cursor()

        stats = {}

        cursor.execute('SELECT COUNT(*), SUM(valor_total) FROM notas_credito')
        row = cursor.fetchone()
        stats['total_notas'] = row[0] or 0
        stats['valor_total'] = row[1] or 0

        cursor.execute('''
            SELECT estado, COUNT(*), SUM(saldo_pendiente)
            FROM notas_credito GROUP BY estado
        ''')
        for row in cursor.fetchall():
            estado_key = row[0].lower()
            stats[f'notas_{estado_key}'] = row[1]
            stats[f'saldo_{estado_key}'] = row[2] or 0

        cursor.execute('SELECT SUM(saldo_pendiente) FROM notas_credito WHERE estado = "PENDIENTE"')
        stats['saldo_pendiente_total'] = cursor.fetchone()[0] or 0

        cursor.execute('SELECT COUNT(*), SUM(valor_aplicado) FROM aplicaciones_notas')
        row = cursor.fetchone()
        stats['total_aplicaciones'] = row[0] or 0
        stats['monto_total_aplicado'] = row[1] or 0

        conn.close()
        _set_cache('notas_stats', stats, 30)
        return jsonify(stats), 200

    except Exception as e:
        logger.error(f"Error en estadisticas_notas: {e}")
        return jsonify({"error": "Error al obtener estadísticas"}), 500


@app.route('/api/notas/por-estado', methods=['GET'])
@jwt_required()
def notas_por_estado():
    try:
        conn = get_db_connection()
        cursor = conn.cursor()

        cursor.execute('''
            SELECT estado, COUNT(*) as cantidad, SUM(valor_total) as valor_total, SUM(saldo_pendiente) as saldo_pendiente
            FROM notas_credito
            GROUP BY estado
        ''')
        resultados = []
        for row in cursor.fetchall():
            resultados.append({
                'estado': row[0],
                'cantidad': row[1] or 0,
                'valor_total': row[2] or 0,
                'saldo_pendiente': row[3] or 0
            })

        conn.close()
        return jsonify(resultados), 200
    except Exception as e:
        logger.error(f"Error en notas_por_estado: {e}")
        return jsonify({"error": "Error al obtener estados"}), 500


@app.route('/api/notas/pendientes', methods=['GET'])
@jwt_required()
def listar_notas_pendientes():
    try:
        estado = request.args.get('estado')
        prioridad = request.args.get('prioridad')
        responsable = request.args.get('responsable')
        fecha_desde = request.args.get('fecha_desde')
        fecha_hasta = request.args.get('fecha_hasta')
        limite = int(request.args.get('limite', 100))
        offset = int(request.args.get('offset', 0))

        query = "SELECT * FROM notas_pendientes WHERE 1=1"
        params = []
        if estado:
            query += " AND estado = ?"
            params.append(estado)
        if prioridad:
            query += " AND prioridad = ?"
            params.append(prioridad)
        if responsable:
            query += " AND responsable = ?"
            params.append(responsable)
        if fecha_desde:
            query += " AND fecha_vencimiento >= ?"
            params.append(fecha_desde)
        if fecha_hasta:
            query += " AND fecha_vencimiento <= ?"
            params.append(fecha_hasta)

        query += " ORDER BY fecha_vencimiento ASC, prioridad DESC LIMIT ? OFFSET ?"
        params.extend([limite, offset])

        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute(query, params)
        items = [dict(row) for row in cursor.fetchall()]

        query_count = query.split('ORDER BY')[0].replace('SELECT *', 'SELECT COUNT(*)')
        cursor.execute(query_count, params[:-2])
        total = cursor.fetchone()[0]
        conn.close()

        return jsonify({
            "items": items,
            "total": total,
            "limite": limite,
            "offset": offset
        }), 200
    except Exception as e:
        logger.error(f"Error en listar_notas_pendientes: {e}")
        return jsonify({"error": "Error al obtener notas pendientes"}), 500


@app.route('/api/notas/pendientes', methods=['POST'])
@jwt_required()
def crear_nota_pendiente():
    try:
        claims, error_response, status_code = _require_write_role()
        if error_response:
            return error_response, status_code
        data = request.get_json() or {}
        numero_nota = data.get('numero_nota')
        if not numero_nota:
            return jsonify({"error": "numero_nota requerido"}), 400

        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute('''
            INSERT INTO notas_pendientes
            (numero_nota, prioridad, fecha_vencimiento, responsable, estado, descripcion)
            VALUES (?, ?, ?, ?, ?, ?)
        ''', (
            numero_nota,
            data.get('prioridad', 'media'),
            data.get('fecha_vencimiento'),
            data.get('responsable'),
            data.get('estado', 'PENDIENTE'),
            data.get('descripcion')
        ))
        conn.commit()
        conn.close()

        registrar_log('notas_pendientes', 'crear', str(numero_nota), claims.get('username'), data)
        return jsonify({"mensaje": "Nota pendiente creada"}), 201
    except Exception as e:
        logger.error(f"Error en crear_nota_pendiente: {e}")
        return jsonify({"error": "Error al crear nota pendiente"}), 500


@app.route('/api/notas/pendientes/<int:nota_id>', methods=['PUT'])
@jwt_required()
def actualizar_nota_pendiente(nota_id):
    try:
        claims, error_response, status_code = _require_write_role()
        if error_response:
            return error_response, status_code
        data = request.get_json() or {}
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute('''
            UPDATE notas_pendientes
            SET numero_nota = ?,
                prioridad = ?,
                fecha_vencimiento = ?,
                responsable = ?,
                estado = ?,
                descripcion = ?,
                fecha_actualizacion = CURRENT_TIMESTAMP
            WHERE id = ?
        ''', (
            data.get('numero_nota'),
            data.get('prioridad'),
            data.get('fecha_vencimiento'),
            data.get('responsable'),
            data.get('estado'),
            data.get('descripcion'),
            nota_id
        ))
        conn.commit()
        conn.close()
        registrar_log('notas_pendientes', 'actualizar', str(nota_id), claims.get('username'), data)
        return jsonify({"mensaje": "Nota pendiente actualizada"}), 200
    except Exception as e:
        logger.error(f"Error en actualizar_nota_pendiente: {e}")
        return jsonify({"error": "Error al actualizar nota pendiente"}), 500


@app.route('/api/notas/pendientes/<int:nota_id>', methods=['DELETE'])
@jwt_required()
def eliminar_nota_pendiente(nota_id):
    try:
        claims, error_response, status_code = _require_write_role()
        if error_response:
            return error_response, status_code
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute('DELETE FROM notas_pendientes WHERE id = ?', (nota_id,))
        conn.commit()
        conn.close()
        registrar_log('notas_pendientes', 'eliminar', str(nota_id), claims.get('username'), {})
        return jsonify({"mensaje": "Nota pendiente eliminada"}), 200
    except Exception as e:
        logger.error(f"Error en eliminar_nota_pendiente: {e}")
        return jsonify({"error": "Error al eliminar nota pendiente"}), 500


@app.route('/api/notas/pendientes/alertas', methods=['GET'])
@jwt_required()
def alertas_notas_pendientes():
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute('''
            SELECT * FROM notas_pendientes
            WHERE estado != 'COMPLETADA'
            ORDER BY fecha_vencimiento ASC
        ''')
        items = [dict(row) for row in cursor.fetchall()]
        conn.close()

        hoy = datetime.now().date()
        vencidas = []
        proximas = []
        for item in items:
            fecha_venc = item.get('fecha_vencimiento')
            if fecha_venc:
                fecha_venc_dt = datetime.fromisoformat(str(fecha_venc)).date()
                if fecha_venc_dt < hoy:
                    vencidas.append(item)
                elif (fecha_venc_dt - hoy).days <= 7:
                    proximas.append(item)

        return jsonify({
            "vencidas": vencidas,
            "proximas": proximas
        }), 200
    except Exception as e:
        logger.error(f"Error en alertas_notas_pendientes: {e}")
        return jsonify({"error": "Error al obtener alertas"}), 500


@app.route('/api/aplicaciones-sistema', methods=['GET'])
@jwt_required()
def listar_aplicaciones_sistema():
    try:
        estado = request.args.get('estado')
        search = request.args.get('search')
        limite = int(request.args.get('limite', 100))
        offset = int(request.args.get('offset', 0))
        query = "SELECT * FROM aplicaciones_sistema WHERE 1=1"
        params = []
        if estado:
            query += " AND estado = ?"
            params.append(estado)
        if search:
            like = f"%{search}%"
            query += " AND (nombre LIKE ? OR version LIKE ?)"
            params.extend([like, like])
        query += " ORDER BY fecha_instalacion DESC LIMIT ? OFFSET ?"
        params.extend([limite, offset])

        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute(query, params)
        items = [dict(row) for row in cursor.fetchall()]
        query_count = query.split('ORDER BY')[0].replace('SELECT *', 'SELECT COUNT(*)')
        cursor.execute(query_count, params[:-2])
        total = cursor.fetchone()[0]
        conn.close()

        return jsonify({
            "items": items,
            "total": total,
            "limite": limite,
            "offset": offset
        }), 200
    except Exception as e:
        logger.error(f"Error en listar_aplicaciones_sistema: {e}")
        return jsonify({"error": "Error al obtener aplicaciones"}), 500


@app.route('/api/aplicaciones-sistema', methods=['POST'])
@jwt_required()
def crear_aplicacion_sistema():
    try:
        claims, error_response, status_code = _require_write_role()
        if error_response:
            return error_response, status_code
        data = request.get_json() or {}
        if not data.get('nombre') or not data.get('version'):
            return jsonify({"error": "nombre y version requeridos"}), 400
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute('''
            INSERT INTO aplicaciones_sistema
            (nombre, version, fecha_instalacion, estado, uso_total)
            VALUES (?, ?, ?, ?, ?)
        ''', (
            data.get('nombre'),
            data.get('version'),
            data.get('fecha_instalacion'),
            data.get('estado', 'ACTIVA'),
            data.get('uso_total', 0)
        ))
        conn.commit()
        conn.close()
        registrar_log('aplicaciones_sistema', 'crear', data.get('nombre'), claims.get('username'), data)
        return jsonify({"mensaje": "Aplicación creada"}), 201
    except Exception as e:
        logger.error(f"Error en crear_aplicacion_sistema: {e}")
        return jsonify({"error": "Error al crear aplicación"}), 500


@app.route('/api/aplicaciones-sistema/<int:app_id>', methods=['PUT'])
@jwt_required()
def actualizar_aplicacion_sistema(app_id):
    try:
        claims, error_response, status_code = _require_write_role()
        if error_response:
            return error_response, status_code
        data = request.get_json() or {}
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute('''
            UPDATE aplicaciones_sistema
            SET nombre = ?,
                version = ?,
                fecha_instalacion = ?,
                estado = ?,
                uso_total = ?,
                fecha_actualizacion = CURRENT_TIMESTAMP
            WHERE id = ?
        ''', (
            data.get('nombre'),
            data.get('version'),
            data.get('fecha_instalacion'),
            data.get('estado'),
            data.get('uso_total'),
            app_id
        ))
        conn.commit()
        conn.close()
        registrar_log('aplicaciones_sistema', 'actualizar', str(app_id), claims.get('username'), data)
        return jsonify({"mensaje": "Aplicación actualizada"}), 200
    except Exception as e:
        logger.error(f"Error en actualizar_aplicacion_sistema: {e}")
        return jsonify({"error": "Error al actualizar aplicación"}), 500


@app.route('/api/aplicaciones-sistema/<int:app_id>', methods=['DELETE'])
@jwt_required()
def eliminar_aplicacion_sistema(app_id):
    try:
        claims, error_response, status_code = _require_write_role()
        if error_response:
            return error_response, status_code
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute('DELETE FROM aplicaciones_sistema WHERE id = ?', (app_id,))
        conn.commit()
        conn.close()
        registrar_log('aplicaciones_sistema', 'eliminar', str(app_id), claims.get('username'), {})
        return jsonify({"mensaje": "Aplicación eliminada"}), 200
    except Exception as e:
        logger.error(f"Error en eliminar_aplicacion_sistema: {e}")
        return jsonify({"error": "Error al eliminar aplicación"}), 500


@app.route('/api/aplicaciones-sistema/<int:app_id>/uso', methods=['POST'])
@jwt_required()
def registrar_uso_aplicacion(app_id):
    try:
        claims, error_response, status_code = _require_write_role()
        if error_response:
            return error_response, status_code
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute('''
            UPDATE aplicaciones_sistema
            SET uso_total = COALESCE(uso_total, 0) + 1,
                ultimo_uso = CURRENT_TIMESTAMP,
                fecha_actualizacion = CURRENT_TIMESTAMP
            WHERE id = ?
        ''', (app_id,))
        conn.commit()
        conn.close()
        return jsonify({"mensaje": "Uso registrado"}), 200
    except Exception as e:
        logger.error(f"Error en registrar_uso_aplicacion: {e}")
        return jsonify({"error": "Error al registrar uso"}), 500


@app.route('/api/aplicaciones/<numero_nota>', methods=['GET'])
@jwt_required()
def obtener_aplicaciones(numero_nota):
    """Obtener aplicaciones de una nota"""
    try:
        conn = get_db_connection()
        cursor = conn.cursor()

        cursor.execute('''
            SELECT * FROM aplicaciones_notas
            WHERE numero_nota = ?
            ORDER BY fecha_aplicacion DESC
        ''', (numero_nota,))

        aplicaciones = [dict(row) for row in cursor.fetchall()]
        conn.close()

        return jsonify(aplicaciones), 200

    except Exception as e:
        logger.error(f"Error en obtener_aplicaciones: {e}")
        return jsonify({"error": "Error al obtener aplicaciones"}), 500


# REPORTE OPERATIVO
@app.route('/api/reporte/operativo', methods=['GET'])
@jwt_required()
def reporte_operativo():
    """Reporte operativo diario"""
    try:
        fecha = request.args.get('fecha')
        if not fecha:
            fecha_obj = datetime.now() - timedelta(days=1)
            fecha = fecha_obj.strftime('%Y-%m-%d')

        conn = get_db_connection()
        cursor = conn.cursor()

        # Notas de la fecha
        cursor.execute('''
            SELECT * FROM notas_credito
            WHERE DATE(fecha_nota) = ? OR DATE(fecha_registro) = ?
            ORDER BY fecha_nota DESC
        ''', (fecha, fecha))
        notas = []
        for row in cursor.fetchall():
            item = dict(row)
            item['nit_cliente'] = _decrypt_value(item.get('nit_encrypted'))
            item['nombre_cliente'] = _decrypt_value(item.get('nombre_cliente_encrypted'))
            notas.append(item)

        # Aplicaciones de la fecha
        cursor.execute('''
            SELECT a.numero_nota, a.numero_factura, a.cantidad_aplicada, a.valor_aplicado, a.fecha_aplicacion,
                   n.nit_encrypted, n.codigo_producto
            FROM aplicaciones_notas a
            LEFT JOIN notas_credito n ON n.id = a.id_nota
            WHERE DATE(a.fecha_aplicacion) = ?
            ORDER BY a.fecha_aplicacion DESC
        ''', (fecha,))
        aplicaciones = []
        for row in cursor.fetchall():
            item = dict(row)
            item['nit_cliente'] = _decrypt_value(item.get('nit_encrypted'))
            aplicaciones.append(item)

        # Facturas rechazadas de la fecha
        cursor.execute('''
            SELECT * FROM facturas_rechazadas
            WHERE DATE(fecha_factura) = ?
            ORDER BY fecha_factura DESC
        ''', (fecha,))
        rechazadas = []
        for row in cursor.fetchall():
            item = dict(row)
            item['nit_cliente'] = _decrypt_value(item.get('nit_encrypted'))
            item['nombre_cliente'] = _decrypt_value(item.get('nombre_cliente_encrypted'))
            rechazadas.append(item)

        # Resumen general
        cursor.execute('''
            SELECT COUNT(*), SUM(saldo_pendiente)
            FROM notas_credito WHERE estado = 'PENDIENTE'
        ''')
        row = cursor.fetchone()
        resumen = {
            'notas_pendientes': row[0] or 0,
            'saldo_pendiente': row[1] or 0
        }

        cursor.execute('SELECT COUNT(*) FROM notas_credito WHERE estado = "APLICADA"')
        resumen['notas_aplicadas'] = cursor.fetchone()[0]

        cursor.execute('SELECT COUNT(*) FROM notas_credito WHERE estado = "NO_APLICADA"')
        resumen['notas_no_aplicadas'] = cursor.fetchone()[0]

        conn.close()

        return jsonify({
            "fecha": fecha,
            "notas_credito": notas,
            "aplicaciones": aplicaciones,
            "facturas_rechazadas": rechazadas,
            "resumen": resumen
        }), 200

    except Exception as e:
        logger.error(f"Error en reporte_operativo: {e}")
        return jsonify({"error": "Error al generar reporte"}), 500


# DASHBOARD
@app.route('/api/dashboard', methods=['GET'])
@jwt_required()
def dashboard():
    """Datos del dashboard principal"""
    try:
        cached = _get_cache('dashboard')
        if cached:
            return jsonify(cached), 200

        conn = get_db_connection()
        cursor = conn.cursor()

        data = {}

        cursor.execute('SELECT COUNT(*), SUM(valor_total) FROM facturas')
        row = cursor.fetchone()
        data['facturas_validas'] = row[0] or 0
        data['valor_total_facturado'] = row[1] or 0

        cursor.execute('SELECT COUNT(*) FROM facturas WHERE registrable = 1')
        data['facturas_registrables'] = cursor.fetchone()[0]

        cursor.execute('SELECT COUNT(*) FROM facturas WHERE registrable = 0')
        data['facturas_no_registrables'] = cursor.fetchone()[0]

        cursor.execute('SELECT COUNT(*), SUM(valor_total) FROM facturas_rechazadas')
        row = cursor.fetchone()
        data['facturas_rechazadas'] = row[0] or 0
        data['valor_rechazado'] = row[1] or 0

        cursor.execute('SELECT COUNT(*), SUM(saldo_pendiente) FROM notas_credito WHERE estado = "PENDIENTE"')
        row = cursor.fetchone()
        data['notas_pendientes'] = row[0] or 0
        data['saldo_pendiente'] = row[1] or 0

        cursor.execute('SELECT COUNT(*) FROM notas_credito WHERE estado = "APLICADA"')
        data['notas_aplicadas'] = cursor.fetchone()[0]

        cursor.execute('SELECT COUNT(*) FROM notas_credito WHERE estado = "NO_APLICADA"')
        data['notas_no_aplicadas'] = cursor.fetchone()[0]

        cursor.execute('''
            SELECT a.numero_nota, a.numero_factura, a.cantidad_aplicada, a.valor_aplicado, a.fecha_aplicacion,
                   n.nit_encrypted, n.codigo_producto
            FROM aplicaciones_notas a
            LEFT JOIN notas_credito n ON n.id = a.id_nota
            ORDER BY a.fecha_aplicacion DESC LIMIT 10
        ''')
        ultimas = []
        for row in cursor.fetchall():
            item = dict(row)
            item['nit_cliente'] = _decrypt_value(item.get('nit_encrypted'))
            ultimas.append(item)
        data['ultimas_aplicaciones'] = ultimas

        conn.close()
        _set_cache('dashboard', data, 30)
        return jsonify(data), 200

    except Exception as e:
        logger.error(f"Error en dashboard: {e}")
        return jsonify({"error": "Error al obtener datos del dashboard"}), 500


# ENDPOINTS DE LOGS
@app.route('/api/admin/logs', methods=['GET'])
@jwt_required()
def listar_logs():
    try:
        claims = get_jwt()
        if claims.get('rol') != 'admin':
            return jsonify({"error": "No tiene permisos para ver logs"}), 403

        entidad = request.args.get('entidad')
        accion = request.args.get('accion')
        usuario = request.args.get('usuario')
        fecha_desde = request.args.get('fecha_desde')
        fecha_hasta = request.args.get('fecha_hasta')
        search = request.args.get('search')
        limite = int(request.args.get('limite', 100))
        offset = int(request.args.get('offset', 0))

        query = "SELECT * FROM audit_logs WHERE 1=1"
        params = []

        if entidad:
            query += " AND entidad = ?"
            params.append(entidad)
        if accion:
            query += " AND accion = ?"
            params.append(accion)
        if usuario:
            query += " AND usuario = ?"
            params.append(usuario)
        if fecha_desde:
            query += " AND fecha_registro >= ?"
            params.append(fecha_desde)
        if fecha_hasta:
            query += " AND fecha_registro <= ?"
            params.append(fecha_hasta)
        if search:
            like = f"%{search}%"
            query += " AND (entidad_id LIKE ? OR payload LIKE ?)"
            params.extend([like, like])

        query += " ORDER BY fecha_registro DESC LIMIT ? OFFSET ?"
        params.extend([limite, offset])

        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute(query, params)
        items = [dict(row) for row in cursor.fetchall()]

        query_count = query.split('ORDER BY')[0].replace('SELECT *', 'SELECT COUNT(*)')
        cursor.execute(query_count, params[:-2])
        total = cursor.fetchone()[0]
        conn.close()

        return jsonify({
            "items": items,
            "total": total,
            "limite": limite,
            "offset": offset
        }), 200
    except Exception as e:
        logger.error(f"Error en listar_logs: {e}")
        return jsonify({"error": "Error al obtener logs"}), 500


# =========================================================================
# ENDPOINTS DE ADMIN - EXPORTACIÓN Y PROCESAMIENTO
# =========================================================================

@app.route('/api/admin/export-preview', methods=['POST'])
@jwt_required()
def exportar_preview_bd():
    try:
        claims = get_jwt()
        if claims.get('rol') != 'admin':
            return jsonify({"error": "No tiene permisos para exportar"}), 403

        data = request.get_json()
        fecha_desde = data.get('fecha_desde')
        fecha_hasta = data.get('fecha_hasta')
        tipo = data.get('tipo', 'facturas')
        limite = int(data.get('limite', 50))

        if not fecha_desde or not fecha_hasta:
            return jsonify({"error": "Fechas requeridas"}), 400

        conn = get_db_connection()
        cursor = conn.cursor()

        if tipo == 'facturas':
            cursor.execute('''
                SELECT numero_linea, numero_factura, codigo_factura, nombre_producto, codigo_producto,
                       nit_encrypted, nombre_cliente_encrypted, cantidad_original, valor_total,
                       cantidad_restante, valor_restante, registrable, total_repeticiones,
                       suma_total_repeticiones, fecha_factura
                FROM facturas
                WHERE fecha_factura >= ? AND fecha_factura <= ?
                ORDER BY fecha_factura DESC, numero_factura
                LIMIT ?
            ''', (fecha_desde, fecha_hasta, limite))
        elif tipo == 'notas':
            cursor.execute('''
                SELECT numero_nota, fecha_nota, nit_encrypted, nombre_cliente_encrypted,
                       codigo_producto, nombre_producto, valor_total, cantidad, saldo_pendiente,
                       cantidad_pendiente, estado
                FROM notas_credito
                WHERE fecha_nota >= ? AND fecha_nota <= ?
                ORDER BY fecha_nota DESC, numero_nota
                LIMIT ?
            ''', (fecha_desde, fecha_hasta, limite))
        elif tipo == 'rechazadas':
            cursor.execute('''
                SELECT numero_factura, fecha_factura, nit_encrypted, nombre_cliente_encrypted,
                       codigo_producto, producto, tipo_inventario, valor_total, razon_rechazo
                FROM facturas_rechazadas
                WHERE fecha_factura >= ? AND fecha_factura <= ?
                ORDER BY fecha_factura DESC, numero_factura
                LIMIT ?
            ''', (fecha_desde, fecha_hasta, limite))
        else:
            cursor.execute('''
                SELECT numero_nota, numero_factura, fecha_factura, nit_hash, codigo_producto,
                       cantidad_aplicada, valor_aplicado, fecha_aplicacion
                FROM aplicaciones_notas
                WHERE fecha_factura >= ? AND fecha_factura <= ?
                ORDER BY fecha_aplicacion DESC
                LIMIT ?
            ''', (fecha_desde, fecha_hasta, limite))

        rows = cursor.fetchall()
        columns = [col[0] for col in cursor.description]
        conn.close()

        usuario = claims.get('username')
        registrar_log('export_preview', 'consultar', f"{tipo}:{fecha_desde}:{fecha_hasta}", usuario, {
            'tipo': tipo,
            'limite': limite
        })

        return jsonify({
            "columnas": columns,
            "rows": [dict(row) for row in rows],
            "limite": limite
        }), 200
    except Exception as e:
        logger.error(f"Error en exportar_preview_bd: {e}")
        return jsonify({"error": "Error al generar preview"}), 500

@app.route('/api/admin/exportar-excel', methods=['POST'])
@jwt_required()
def exportar_excel_bd():
    """
    Exporta datos de la BD a Excel por rango de fechas
    Solo para admins
    """
    try:
        claims = get_jwt()
        if claims.get('rol') != 'admin':
            return jsonify({"error": "No tiene permisos para exportar"}), 403

        data = request.get_json()
        fecha_desde = data.get('fecha_desde')
        fecha_hasta = data.get('fecha_hasta')
        tipo = data.get('tipo', 'facturas')  # facturas, notas, rechazadas

        if not fecha_desde or not fecha_hasta:
            return jsonify({"error": "Fechas requeridas"}), 400

        conn = get_db_connection()
        cursor = conn.cursor()

        rows_data = []
        if tipo == 'facturas':
            cursor.execute('''
                SELECT numero_linea, numero_factura, codigo_factura, nombre_producto, codigo_producto,
                       nit_encrypted, nombre_cliente_encrypted, cantidad_original, valor_total,
                       cantidad_restante, valor_restante, registrable, total_repeticiones,
                       suma_total_repeticiones, fecha_factura
                FROM facturas
                WHERE fecha_factura >= ? AND fecha_factura <= ?
                ORDER BY fecha_factura DESC, numero_factura
            ''', (fecha_desde, fecha_hasta))
            columns = ['Linea', 'Factura', 'Codigo Factura', 'Producto', 'Codigo Producto', 'Cliente', 'NIT',
                      'Cantidad', 'Valor Total', 'Cant Rest', 'Valor Rest', 'Registrable',
                      'Repeticiones', 'Suma Repeticiones', 'Fecha']
            for row in cursor.fetchall():
                item = list(row)
                item[5] = _decrypt_value(item[5])
                item[6] = _decrypt_value(item[6])
                rows_data.append(item)

        elif tipo == 'notas':
            cursor.execute('''
                SELECT numero_nota, fecha_nota, nombre_cliente_encrypted, nit_encrypted,
                       nombre_producto, codigo_producto, cantidad, valor_total,
                       cantidad_pendiente, saldo_pendiente, estado, es_agente, causal_devolucion
                FROM notas_credito
                WHERE fecha_nota >= ? AND fecha_nota <= ?
                ORDER BY fecha_nota DESC
            ''', (fecha_desde, fecha_hasta))
            columns = ['Nota', 'Fecha', 'Cliente', 'NIT', 'Producto', 'Codigo',
                      'Cantidad', 'Valor Total', 'Cant Pend', 'Saldo Pend',
                      'Estado', 'Es Agente', 'Causal']
            for row in cursor.fetchall():
                item = list(row)
                item[2] = _decrypt_value(item[2])
                item[3] = _decrypt_value(item[3])
                rows_data.append(item)

        elif tipo == 'rechazadas':
            cursor.execute('''
                SELECT numero_factura, numero_linea, producto, codigo_producto,
                       nombre_cliente_encrypted, nit_encrypted, cantidad, valor_total,
                       tipo_inventario, razon_rechazo, fecha_factura
                FROM facturas_rechazadas
                WHERE fecha_factura >= ? AND fecha_factura <= ?
                ORDER BY fecha_factura DESC
            ''', (fecha_desde, fecha_hasta))
            columns = ['Factura', 'Linea', 'Producto', 'Codigo', 'Cliente', 'NIT',
                      'Cantidad', 'Valor', 'Tipo Inv', 'Razon Rechazo', 'Fecha']
            for row in cursor.fetchall():
                item = list(row)
                item[4] = _decrypt_value(item[4])
                item[5] = _decrypt_value(item[5])
                rows_data.append(item)

        elif tipo == 'aplicaciones':
            cursor.execute('''
                SELECT a.numero_nota, a.numero_factura, a.numero_linea, n.nit_encrypted,
                       a.codigo_producto, a.cantidad_aplicada, a.valor_aplicado, a.fecha_aplicacion
                FROM aplicaciones_notas
                LEFT JOIN notas_credito n ON n.id = a.id_nota
                WHERE DATE(a.fecha_aplicacion) >= ? AND DATE(a.fecha_aplicacion) <= ?
                ORDER BY a.fecha_aplicacion DESC
            ''', (fecha_desde, fecha_hasta))
            columns = ['Nota', 'Factura', 'Linea', 'NIT', 'Codigo',
                      'Cantidad Aplicada', 'Valor Aplicado', 'Fecha']
            for row in cursor.fetchall():
                item = list(row)
                item[3] = _decrypt_value(item[3])
                rows_data.append(item)
        else:
            return jsonify({"error": "Tipo inválido"}), 400

        conn.close()

        if not rows_data:
            return jsonify({"error": "No hay datos para el rango seleccionado"}), 404

        # Generar Excel
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
        except ImportError:
            return jsonify({"error": "openpyxl no instalado"}), 500

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = tipo.capitalize()

        # Estilos
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Headers
        for col, header in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border

        # Datos
        for row_idx, row in enumerate(rows_data, 2):
            for col_idx, value in enumerate(row, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = thin_border

        # Ajustar anchos
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column].width = adjusted_width

        # Guardar
        output_dir = PROJECT_ROOT / 'output'
        output_dir.mkdir(exist_ok=True)

        filename = f"export_{tipo}_{fecha_desde}_{fecha_hasta}.xlsx"
        output_path = output_dir / filename

        wb.save(str(output_path))
        usuario = claims.get('username')
        registrar_log('export_excel', 'crear', filename, usuario, {
            'tipo': tipo,
            'fecha_desde': fecha_desde,
            'fecha_hasta': fecha_hasta,
            'total_registros': len(rows_data)
        })

        return jsonify({
            "exito": True,
            "mensaje": f"Excel generado con {len(rows_data)} registros",
            "archivo": filename,
            "total_registros": len(rows_data)
        }), 200

    except Exception as e:
        logger.error(f"Error en exportar_excel_bd: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({"error": f"Error al exportar: {str(e)}"}), 500


@app.route('/api/admin/exportar-pdf', methods=['POST'])
@jwt_required()
def exportar_pdf_bd():
    try:
        claims = get_jwt()
        if claims.get('rol') != 'admin':
            return jsonify({"error": "No tiene permisos para exportar"}), 403

        data = request.get_json()
        fecha_desde = data.get('fecha_desde')
        fecha_hasta = data.get('fecha_hasta')
        tipo = data.get('tipo', 'facturas')

        if not fecha_desde or not fecha_hasta:
            return jsonify({"error": "Fechas requeridas"}), 400

        conn = get_db_connection()
        cursor = conn.cursor()

        if tipo == 'facturas':
            cursor.execute('''
                SELECT numero_factura, codigo_factura, nombre_producto, codigo_producto,
                       nit_encrypted, cantidad_original, valor_total, fecha_factura
                FROM facturas
                WHERE fecha_factura >= ? AND fecha_factura <= ?
                ORDER BY fecha_factura DESC, numero_factura
            ''', (fecha_desde, fecha_hasta))
            columns = ['Factura', 'Codigo', 'Producto', 'Cod Prod', 'NIT', 'Cantidad', 'Valor', 'Fecha']
        elif tipo == 'notas':
            cursor.execute('''
                SELECT numero_nota, nombre_producto, codigo_producto, nit_encrypted,
                       cantidad, valor_total, estado, fecha_nota
                FROM notas_credito
                WHERE fecha_nota >= ? AND fecha_nota <= ?
                ORDER BY fecha_nota DESC
            ''', (fecha_desde, fecha_hasta))
            columns = ['Nota', 'Producto', 'Codigo', 'NIT', 'Cantidad', 'Valor', 'Estado', 'Fecha']
        else:
            conn.close()
            return jsonify({"error": "Tipo inválido"}), 400

        rows = []
        for row in cursor.fetchall():
            item = list(row)
            if tipo == 'facturas':
                item[4] = _decrypt_value(item[4])
            elif tipo == 'notas':
                item[3] = _decrypt_value(item[3])
            rows.append(item)
        conn.close()

        if not rows:
            return jsonify({"error": "No hay datos para el rango seleccionado"}), 404

        try:
            from reportlab.lib.pagesizes import letter
            from reportlab.pdfgen import canvas
        except ImportError:
            return jsonify({"error": "reportlab no instalado"}), 500

        output_dir = PROJECT_ROOT / 'output'
        output_dir.mkdir(exist_ok=True)

        filename = f"export_{tipo}_{fecha_desde}_{fecha_hasta}.pdf"
        output_path = output_dir / filename

        c = canvas.Canvas(str(output_path), pagesize=letter)
        width, height = letter
        y = height - 40

        c.setFont("Helvetica-Bold", 12)
        c.drawString(40, y, f"Reporte {tipo} {fecha_desde} a {fecha_hasta}")
        y -= 20

        c.setFont("Helvetica-Bold", 9)
        c.drawString(40, y, " | ".join(columns))
        y -= 14

        c.setFont("Helvetica", 8)
        for row in rows:
            line = " | ".join([str(v) for v in row])
            if y < 40:
                c.showPage()
                y = height - 40
                c.setFont("Helvetica-Bold", 9)
                c.drawString(40, y, " | ".join(columns))
                y -= 14
                c.setFont("Helvetica", 8)
            c.drawString(40, y, line[:200])
            y -= 12

        c.save()
        usuario = claims.get('username')
        registrar_log('export_pdf', 'crear', filename, usuario, {
            'tipo': tipo,
            'fecha_desde': fecha_desde,
            'fecha_hasta': fecha_hasta,
            'total_registros': len(rows)
        })

        return jsonify({
            "exito": True,
            "mensaje": f"PDF generado con {len(rows)} registros",
            "archivo": filename,
            "total_registros": len(rows)
        }), 200

    except Exception as e:
        logger.error(f"Error en exportar_pdf_bd: {e}")
        return jsonify({"error": "Error al exportar PDF"}), 500


@app.route('/api/admin/descargar/<filename>', methods=['GET'])
@jwt_required()
def descargar_archivo(filename):
    """Descarga un archivo generado"""
    try:
        from flask import send_file

        # Validar nombre de archivo (seguridad)
        if '..' in filename or '/' in filename:
            return jsonify({"error": "Nombre de archivo inválido"}), 400

        output_dir = PROJECT_ROOT / 'output'
        file_path = output_dir / filename

        if not file_path.exists():
            return jsonify({"error": "Archivo no encontrado"}), 404

        return send_file(
            str(file_path),
            as_attachment=True,
            download_name=filename
        )

    except Exception as e:
        logger.error(f"Error al descargar archivo: {e}")
        return jsonify({"error": "Error al descargar archivo"}), 500


@app.route('/api/admin/procesar-rango', methods=['POST'])
@jwt_required()
def procesar_rango():
    """
    Procesa un rango de fechas desde la API externa y guarda en BD
    Solo para admins - Proceso largo
    """
    try:
        claims = get_jwt()
        if claims.get('rol') not in ('admin', 'editor'):
            return jsonify({"error": "No tiene permisos"}), 403

        data = request.get_json()
        fecha_desde_str = data.get('fecha_desde')
        fecha_hasta_str = data.get('fecha_hasta')

        if not fecha_desde_str or not fecha_hasta_str:
            return jsonify({"error": "Fechas requeridas"}), 400

        fecha_desde = datetime.strptime(fecha_desde_str, '%Y-%m-%d')
        fecha_hasta = datetime.strptime(fecha_hasta_str, '%Y-%m-%d')

        # Validar rango máximo 90 días
        diff_days = (fecha_hasta - fecha_desde).days
        if diff_days > 90:
            return jsonify({"error": "Rango máximo permitido: 90 días"}), 400

        if fecha_desde > fecha_hasta:
            return jsonify({"error": "Fecha desde debe ser anterior a fecha hasta"}), 400

        # Importar función de procesamiento
        sys.path.insert(0, str(BACKEND_DIR))
        from main import procesar_rango_fechas

        # Configuración
        config = {
            'CONNI_KEY': os.getenv('CONNI_KEY'),
            'CONNI_TOKEN': os.getenv('CONNI_TOKEN'),
            'TEMPLATE_PATH': os.getenv('TEMPLATE_PATH', './templates/plantilla.xlsx')
        }

        if not config['CONNI_KEY'] or not config['CONNI_TOKEN']:
            return jsonify({"error": "Credenciales API no configuradas"}), 500

        # Ejecutar procesamiento
        resultado = procesar_rango_fechas(fecha_desde, fecha_hasta, config)

        return jsonify(resultado), 200

    except Exception as e:
        logger.error(f"Error en procesar_rango: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({"error": f"Error al procesar: {str(e)}"}), 500


@app.route('/api/admin/archivos', methods=['GET'])
@jwt_required()
def listar_archivos():
    """Lista archivos disponibles para descarga"""
    try:
        output_dir = PROJECT_ROOT / 'output'

        if not output_dir.exists():
            return jsonify({"archivos": []}), 200

        archivos = []
        for f in output_dir.iterdir():
            if f.is_file() and f.suffix in ['.xlsx', '.txt', '.pdf']:
                archivos.append({
                    'nombre': f.name,
                    'tamaño': f.stat().st_size,
                    'fecha_modificacion': datetime.fromtimestamp(f.stat().st_mtime).isoformat()
                })

        archivos.sort(key=lambda x: x['fecha_modificacion'], reverse=True)

        return jsonify({"archivos": archivos}), 200

    except Exception as e:
        logger.error(f"Error al listar archivos: {e}")
        return jsonify({"error": "Error al listar archivos"}), 500


@app.route('/api/openapi.json', methods=['GET'])
def openapi_spec():
    spec = {
        "openapi": "3.0.0",
        "info": {
            "title": "CIPA Notas Crédito y Facturas API",
            "version": "2.0.0"
        },
        "paths": {
            "/api/facturas": {
                "get": {"summary": "Listar facturas"},
                "post": {"summary": "Crear factura"}
            },
            "/api/facturas/{factura_id}": {
                "get": {"summary": "Obtener factura"},
                "put": {"summary": "Actualizar factura"},
                "delete": {"summary": "Eliminar factura"}
            },
            "/api/notas": {
                "get": {"summary": "Listar notas"},
                "post": {"summary": "Crear nota"}
            },
            "/api/notas/{nota_id}": {
                "get": {"summary": "Obtener nota"},
                "put": {"summary": "Actualizar nota"},
                "delete": {"summary": "Eliminar nota"}
            },
            "/api/notas/aplicar": {
                "post": {"summary": "Aplicar nota a factura"}
            },
            "/api/notas/no-aplicadas": {
                "get": {"summary": "Listar motivos de no aplicación"},
                "post": {"summary": "Registrar motivo de no aplicación"}
            },
            "/api/notas/estadisticas": {
                "get": {"summary": "Estadísticas de notas"}
            },
            "/api/notas/por-estado": {
                "get": {"summary": "Notas por estado"}
            },
            "/api/facturas/estadisticas": {
                "get": {"summary": "Estadísticas de facturas"}
            },
            "/api/facturas/transacciones": {
                "get": {"summary": "Transacciones recientes"}
            },
            "/api/dashboard": {
                "get": {"summary": "Dashboard"}
            },
            "/api/admin/exportar-excel": {
                "post": {"summary": "Exportar Excel"}
            },
            "/api/admin/exportar-pdf": {
                "post": {"summary": "Exportar PDF"}
            }
        }
    }
    return jsonify(spec), 200


# HEALTH CHECK
@app.route('/api/health', methods=['GET'])
def health():
    """Health check - sin autenticación"""
    return jsonify({
        "status": "healthy",
        "timestamp": datetime.now().isoformat(),
        "version": "2.0.0"
    }), 200


# ERROR HANDLERS
@app.errorhandler(404)
def not_found(error):
    return jsonify({"error": "Endpoint no encontrado"}), 404


@app.errorhandler(500)
def internal_error(error):
    return jsonify({"error": "Error interno del servidor"}), 500


if __name__ == '__main__':
    port = int(os.getenv('PORT', os.getenv('API_PORT', 2500)))
    debug_enabled = os.getenv('FLASK_DEBUG', 'false').lower() == 'true'
    logger.info(f"Iniciando API en puerto {port}")
    engine = get_engine()
    if engine == 'mysql':
        logger.info("Base de datos: MySQL")
    else:
        logger.info(f"Base de datos: {DB_PATH}")
    app.run(host='0.0.0.0', port=port, debug=debug_enabled)
